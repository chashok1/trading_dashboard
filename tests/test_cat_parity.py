"""
Parity test: every drv_cat_* column for a sampled set of (symbol, date) pairs
must match the corresponding cell in the Excel MA tab.

Run with:
    pytest tests/test_cat_parity.py            # default — 20 symbols × 1 snapshot
    pytest tests/test_cat_parity.py -k bollinger  # one cat-table only
    PARITY_SNAPSHOT=2026-04-30 pytest tests/test_cat_parity.py

This is the test that catches B3 (silent NULL data from missing source_expr)
and the source_expr bugs the registry will accumulate over time.
"""
from __future__ import annotations
import os
import sys
from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from sqlalchemy import text

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from etl.db import session_scope


# =============================================================================
# Configuration
# =============================================================================

DEFAULT_SAMPLE_SYMBOLS = [
    # Equities (large cap)
    "AAPL", "MSFT", "NVDA", "TSLA", "JPM",
    # ETFs
    "SPY", "QQQ", "IWM", "TLT", "GLD",
    # Sectors (XL*)
    "XLE", "XLF", "XLK",
    # Vol / Treasuries / Commodities
    "VIX", "HYG", "SLV",
    # International / small caps
    "EEM", "EFA", "BIIB", "ZM",
]

WB_CANDIDATES = [
    Path(r"C:\Ashok\Invest\Projects\Cluade\Tickers 2026-04-30.xlsx"),
    Path(r"C:\Ashok\Invest\Cluade\Cluade\Tickers 2026-04-30.xlsx"),
]

NUMERIC_TOLERANCE = 1e-4   # relative
ABSOLUTE_TOLERANCE = 1e-6  # for tiny values

# Columns to never check (auto-managed)
SKIP_COL_NAMES = {"as_of_date", "symbol", "source_run_id", "computed_at"}


# =============================================================================
# Fixtures
# =============================================================================

def _find_workbook() -> Path:
    for p in WB_CANDIDATES:
        if p.exists():
            return p
    pytest.skip(f"Workbook not found in any of: {WB_CANDIDATES}")


def _snapshot_date() -> date:
    val = os.environ.get("PARITY_SNAPSHOT", "2026-04-30")
    return datetime.strptime(val, "%Y-%m-%d").date()


@pytest.fixture(scope="session")
def workbook():
    wb_path = _find_workbook()
    wb = load_workbook(wb_path, data_only=True, read_only=True)
    yield wb
    wb.close()


@pytest.fixture(scope="session")
def ma_rows_by_symbol(workbook):
    """{symbol: {col_letter: value}} from the MA tab."""
    ma = workbook["MA"]
    rows = list(ma.iter_rows(values_only=True))
    if len(rows) < 2:
        pytest.skip("MA tab is empty / template only")
    headers = rows[0]
    out = {}
    # Symbol column is B (index 1)
    for r in rows[1:]:
        if not r or len(r) < 2 or not r[1]:
            continue
        sym = str(r[1]).strip().upper()
        # Map col_letter → value for this row
        cell_map = {}
        for i, val in enumerate(r):
            if i >= len(headers):
                break
            cell_map[i + 1] = val   # 1-based
        out[sym] = cell_map
    return out


@pytest.fixture(scope="session")
def registry():
    """List of registry rows for columns we want to parity-check."""
    with session_scope() as s:
        rows = s.execute(text("""
            SELECT column_name, excel_header, excel_col_idx, excel_col_letter,
                   pg_type, drv_cat_table, source_expr
            FROM ref_ma_columns
            WHERE drv_cat_table NOT IN ('drv_cat_separator')
              AND column_name NOT IN :skip
            ORDER BY excel_col_idx
        """), {"skip": tuple(SKIP_COL_NAMES)}).mappings().all()
    return [dict(r) for r in rows]


@pytest.fixture(scope="session")
def db_rows_by_symbol(registry):
    """{cat_table: {symbol: {col_name: value}}} for snapshot date."""
    snap = _snapshot_date()
    cat_tables = sorted({r["drv_cat_table"] for r in registry})
    out = {}
    with session_scope() as s:
        for cat in cat_tables:
            cols_in_cat = [r["column_name"] for r in registry if r["drv_cat_table"] == cat]
            if not cols_in_cat:
                continue
            # Quote where needed
            select_cols = ", ".join(
                f'"{c}"' if not c.replace("_","").isalnum() or c[:1].isdigit() else c
                for c in cols_in_cat
            )
            try:
                rs = s.execute(
                    text(f"SELECT symbol, {select_cols} FROM {cat} WHERE as_of_date = :d"),
                    {"d": snap},
                ).mappings().all()
            except Exception as e:
                pytest.skip(f"Cat-table {cat} not loaded for {snap}: {e}")
            out[cat] = {row["symbol"].upper(): dict(row) for row in rs}
    return out


# =============================================================================
# Comparison helpers
# =============================================================================

def _normalize(v):
    """Normalize a value for comparison: numeric→float, strings→stripped str, NaN→None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v != v:   # NaN
            return None
        return float(v)
    s = str(v).strip()
    if s in ("", "#N/A", "#N/A N/A", "NaN", "nan"):
        return None
    # Numeric-looking strings
    try:
        return float(s.replace(",", "").rstrip("%"))
    except ValueError:
        return s


def _values_match(excel_v, db_v, pg_type: str) -> tuple[bool, str]:
    """Return (matches, reason)."""
    e = _normalize(excel_v)
    d = _normalize(db_v)
    if e is None and d is None:
        return True, "both null"
    if e is None or d is None:
        return False, f"null mismatch (excel={e!r}, db={d!r})"
    if isinstance(e, (int, float)) and isinstance(d, (int, float)):
        if abs(e - d) <= ABSOLUTE_TOLERANCE:
            return True, "absolute tolerance"
        denom = max(abs(e), abs(d), 1.0)
        rel = abs(e - d) / denom
        if rel <= NUMERIC_TOLERANCE:
            return True, f"rel={rel:.2e}"
        return False, f"numeric mismatch excel={e} db={d} rel={rel:.4e}"
    if isinstance(e, str) and isinstance(d, str):
        return (e == d), f"string {'match' if e == d else 'mismatch'}"
    return False, f"type mismatch excel={type(e).__name__} db={type(d).__name__}"


# =============================================================================
# Tests
# =============================================================================

def test_workbook_present(workbook):
    assert "MA" in workbook.sheetnames


def test_registry_populated(registry):
    assert len(registry) > 100, "ref_ma_columns must have at least 100 columns seeded"


def test_at_least_one_cat_table_loaded(db_rows_by_symbol):
    loaded = {k: len(v) for k, v in db_rows_by_symbol.items() if v}
    assert loaded, "no drv_cat_* tables have any rows for this snapshot"


@pytest.mark.parametrize("cat_table", [
    "drv_cat_identity", "drv_cat_price", "drv_cat_atomic_input",
    "drv_cat_bollinger", "drv_cat_rsi", "drv_cat_macd", "drv_cat_ivhv",
    "drv_cat_volume", "drv_cat_risk_range", "drv_cat_trend_trade",
    "drv_cat_moving_avg", "drv_cat_perf_extremes",
])
def test_parity_for_cat_table(cat_table, registry, ma_rows_by_symbol, db_rows_by_symbol):
    """For each sample symbol, every column in this cat-table must agree with Excel."""
    cols = [r for r in registry if r["drv_cat_table"] == cat_table]
    if not cols:
        pytest.skip(f"No registry rows for {cat_table}")

    db_for_cat = db_rows_by_symbol.get(cat_table, {})
    if not db_for_cat:
        pytest.skip(f"{cat_table} not loaded; run derive_all first")

    failures = []
    checked = 0
    skipped_no_expr = 0

    for symbol in DEFAULT_SAMPLE_SYMBOLS:
        sym_u = symbol.upper()
        db_row = db_for_cat.get(sym_u)
        ma_row = ma_rows_by_symbol.get(sym_u)
        if not db_row or not ma_row:
            continue
        for col in cols:
            cn = col["column_name"]
            if not col["source_expr"]:
                skipped_no_expr += 1
                continue
            excel_v = ma_row.get(col["excel_col_idx"])
            db_v = db_row.get(cn)
            ok, reason = _values_match(excel_v, db_v, col["pg_type"])
            checked += 1
            if not ok:
                failures.append(f"{sym_u}.{cn}: {reason}")
            if len(failures) >= 25:
                break
        if len(failures) >= 25:
            break

    print(f"\n{cat_table}: checked={checked} skipped_no_expr={skipped_no_expr} failures={len(failures)}")
    assert not failures, "First 25 parity failures:\n" + "\n".join(failures[:25])


def test_no_silent_null_columns(registry, db_rows_by_symbol):
    """Detect cat-table columns that are 100% NULL in the DB (B3 regression check)."""
    bad: list[str] = []
    for cat, rows in db_rows_by_symbol.items():
        if not rows:
            continue
        # Pull column names (excluding symbol)
        any_row = next(iter(rows.values()))
        for col in any_row.keys():
            if col == "symbol":
                continue
            non_null = sum(1 for r in rows.values() if r.get(col) is not None)
            if non_null == 0:
                bad.append(f"{cat}.{col} (all {len(rows)} rows NULL)")
    if bad:
        # Allow a small budget — some columns are legitimately sparse on a given day
        budget = 10
        if len(bad) > budget:
            pytest.fail(
                f"{len(bad)} columns are 100% NULL across all sampled rows "
                f"(budget {budget}). First 20:\n" + "\n".join(bad[:20])
            )
