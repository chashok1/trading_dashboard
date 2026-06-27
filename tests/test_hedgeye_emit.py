"""
Tests for etl/hedgeye/emit.py — pure renderers (no DB, no network).

Each renderer takes parsed rows (dicts from parsers.py) and an output path,
and writes the file in the exact format the existing loader understands.
These tests verify sheet name, headers, row count, and selected values.
"""
from __future__ import annotations

import csv
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from etl.hedgeye.emit import (
    render_etf_changes,
    render_investing_ideas,
    render_portfolio_solutions,
    render_risk_range,
    render_the_call,
)


# ── helpers ────────────────────────────────────────────────────────────────────

def _rr_rows(n: int = 3) -> list[dict]:
    return [
        {
            "snapshot_date": date(2026, 6, 26),
            "market_close": date(2026, 6, 25),
            "symbol": f"SYM{i}",
            "tos_symbol": f"SYM{i}",
            "name": f"Name {i}",
            "outlook": "BULLISH" if i % 2 == 0 else "BEARISH",
            "buy_trade": 100.0 + i,
            "sell_trade": 110.0 + i,
            "last_price": 105.0 + i,
        }
        for i in range(n)
    ]


def _iichg_rows(n: int = 2) -> list[dict]:
    return [
        {
            "snapshot_date": date(2026, 6, 26),
            "message_id": "<msg@h>",
            "action": "add" if i % 2 == 0 else "remove",
            "side": "long",
            "symbol": f"TICK{i}",
            "tos_symbol": f"TICK{i}",
        }
        for i in range(n)
    ]


def _etfchg_rows(n: int = 2) -> list[dict]:
    return [
        {
            "snapshot_date": date(2026, 6, 26),
            "message_id": "<msg@h>",
            "action": "add",
            "side": "long",
            "symbol": f"ETF{i}",
            "tos_symbol": f"ETF{i}",
        }
        for i in range(n)
    ]


def _ps_rows(n: int = 3) -> list[dict]:
    return [
        {
            "snapshot_date": date(2026, 6, 26),
            "message_id": "<msg@h>",
            "rank": i + 1,
            "ticker": f"PS{i}",
            "tos_symbol": f"PS{i}",
        }
        for i in range(n)
    ]


def _call_rows(n: int = 3) -> list[dict]:
    return [
        {
            "snapshot_date": date(2026, 6, 26),
            "message_id": "<msg@h>",
            "symbol": f"CALL{i}",
            "tos_symbol": f"CALL{i}",
            "outlook": "long" if i == 0 else "short",
        }
        for i in range(n)
    ]


# ── render_risk_range ─────────────────────────────────────────────────────────

class TestRenderRiskRange:
    def test_creates_file(self, tmp_path):
        p = tmp_path / "RR 2026-06-26.xlsx"
        render_risk_range(_rr_rows(), p)
        assert p.exists()

    def test_sheet_name(self, tmp_path):
        p = tmp_path / "RR 2026-06-26.xlsx"
        render_risk_range(_rr_rows(), p)
        wb = openpyxl.load_workbook(str(p), data_only=True)
        assert "Table_Section" in wb.sheetnames

    def test_headers(self, tmp_path):
        p = tmp_path / "RR 2026-06-26.xlsx"
        render_risk_range(_rr_rows(), p)
        wb = openpyxl.load_workbook(str(p), data_only=True)
        ws = wb["Table_Section"]
        headers = [ws.cell(1, c).value for c in range(1, 8)]
        assert headers == [
            "Index", "Description", "Outlook",
            "BUY TRADE", "SELL TRADE", "Prev Close", "RR Date",
        ]

    def test_row_count(self, tmp_path):
        p = tmp_path / "RR 2026-06-26.xlsx"
        rows = _rr_rows(5)
        render_risk_range(rows, p)
        wb = openpyxl.load_workbook(str(p), data_only=True)
        ws = wb["Table_Section"]
        # header + 5 data rows
        assert ws.max_row == 6

    def test_symbol_in_index_col(self, tmp_path):
        p = tmp_path / "RR 2026-06-26.xlsx"
        render_risk_range(_rr_rows(1), p)
        wb = openpyxl.load_workbook(str(p), data_only=True)
        ws = wb["Table_Section"]
        assert ws.cell(2, 1).value == "SYM0"

    def test_buy_trade_numeric(self, tmp_path):
        p = tmp_path / "RR 2026-06-26.xlsx"
        render_risk_range(_rr_rows(1), p)
        wb = openpyxl.load_workbook(str(p), data_only=True)
        ws = wb["Table_Section"]
        assert ws.cell(2, 4).value == 100.0

    def test_empty_rows_creates_header_only(self, tmp_path):
        p = tmp_path / "RR 2026-06-26.xlsx"
        render_risk_range([], p)
        wb = openpyxl.load_workbook(str(p), data_only=True)
        ws = wb["Table_Section"]
        assert ws.max_row == 1


# ── render_investing_ideas ────────────────────────────────────────────────────

class TestRenderInvestingIdeas:
    def test_creates_file(self, tmp_path):
        p = tmp_path / "IIChange 2026-06-26.xlsx"
        render_investing_ideas(_iichg_rows(), p)
        assert p.exists()

    def test_sheet_name(self, tmp_path):
        """Sheet name is 'Data Sheet' to mirror ETFChange (design decision 2026-06-27)."""
        p = tmp_path / "IIChange 2026-06-26.xlsx"
        render_investing_ideas(_iichg_rows(), p)
        wb = openpyxl.load_workbook(str(p), data_only=True)
        assert "Data Sheet" in wb.sheetnames

    def test_headers(self, tmp_path):
        """Headers mirror ETFChange exactly: leading spaces on cols 2-5."""
        p = tmp_path / "IIChange 2026-06-26.xlsx"
        render_investing_ideas(_iichg_rows(), p)
        wb = openpyxl.load_workbook(str(p), data_only=True)
        ws = wb["Data Sheet"]
        headers = [ws.cell(1, c).value for c in range(1, 6)]
        assert headers == ["Date", " Description", " Ticker", " Outlook", " Action"]

    def test_row_count(self, tmp_path):
        p = tmp_path / "IIChange 2026-06-26.xlsx"
        render_investing_ideas(_iichg_rows(4), p)
        wb = openpyxl.load_workbook(str(p), data_only=True)
        ws = wb["Data Sheet"]
        assert ws.max_row == 5  # header + 4 rows

    def test_action_col_contains_add_remove(self, tmp_path):
        """Action is col 5 (' Action'), same position as ETFChange."""
        p = tmp_path / "IIChange 2026-06-26.xlsx"
        render_investing_ideas(_iichg_rows(2), p)
        wb = openpyxl.load_workbook(str(p), data_only=True)
        ws = wb["Data Sheet"]
        actions = [ws.cell(r, 5).value for r in range(2, 4)]
        assert set(actions) == {"add", "remove"}

    def test_ticker_col_populated(self, tmp_path):
        """Ticker is col 3 (' Ticker'), same position as ETFChange."""
        p = tmp_path / "IIChange 2026-06-26.xlsx"
        render_investing_ideas(_iichg_rows(1), p)
        wb = openpyxl.load_workbook(str(p), data_only=True)
        ws = wb["Data Sheet"]
        assert ws.cell(2, 3).value == "TICK0"


# ── render_etf_changes ────────────────────────────────────────────────────────

class TestRenderEtfChanges:
    def test_creates_file(self, tmp_path):
        p = tmp_path / "ETFChange 2026-06-26.xlsx"
        render_etf_changes(_etfchg_rows(), p)
        assert p.exists()

    def test_sheet_name(self, tmp_path):
        p = tmp_path / "ETFChange 2026-06-26.xlsx"
        render_etf_changes(_etfchg_rows(), p)
        wb = openpyxl.load_workbook(str(p), data_only=True)
        assert "Data Sheet" in wb.sheetnames

    def test_headers_with_leading_spaces(self, tmp_path):
        """Leading spaces on cols 2-5 must be preserved for Excel parity."""
        p = tmp_path / "ETFChange 2026-06-26.xlsx"
        render_etf_changes(_etfchg_rows(), p)
        wb = openpyxl.load_workbook(str(p), data_only=True)
        ws = wb["Data Sheet"]
        headers = [ws.cell(1, c).value for c in range(1, 6)]
        assert headers == ["Date", " Description", " Ticker", " Outlook", " Action"]

    def test_row_count(self, tmp_path):
        p = tmp_path / "ETFChange 2026-06-26.xlsx"
        render_etf_changes(_etfchg_rows(3), p)
        wb = openpyxl.load_workbook(str(p), data_only=True)
        ws = wb["Data Sheet"]
        assert ws.max_row == 4  # header + 3 rows

    def test_ticker_in_col3(self, tmp_path):
        """load_etfchg reads ticker from column 3 (position-based)."""
        p = tmp_path / "ETFChange 2026-06-26.xlsx"
        render_etf_changes(_etfchg_rows(1), p)
        wb = openpyxl.load_workbook(str(p), data_only=True)
        ws = wb["Data Sheet"]
        assert ws.cell(2, 3).value == "ETF0"

    def test_action_in_col5(self, tmp_path):
        """load_etfchg reads action from column 5 (position-based)."""
        p = tmp_path / "ETFChange 2026-06-26.xlsx"
        render_etf_changes(_etfchg_rows(1), p)
        wb = openpyxl.load_workbook(str(p), data_only=True)
        ws = wb["Data Sheet"]
        assert ws.cell(2, 5).value == "add"

    def test_exactly_5_columns(self, tmp_path):
        """load_etfchg new-format detection: max_column <= 5."""
        p = tmp_path / "ETFChange 2026-06-26.xlsx"
        render_etf_changes(_etfchg_rows(2), p)
        wb = openpyxl.load_workbook(str(p), data_only=True)
        ws = wb["Data Sheet"]
        assert ws.max_column == 5


# ── render_portfolio_solutions ────────────────────────────────────────────────

class TestRenderPortfolioSolutions:
    def test_creates_file(self, tmp_path):
        p = tmp_path / "PS 2026-06-26.xlsx"
        render_portfolio_solutions(_ps_rows(), p)
        assert p.exists()

    def test_sheet_name(self, tmp_path):
        p = tmp_path / "PS 2026-06-26.xlsx"
        render_portfolio_solutions(_ps_rows(), p)
        wb = openpyxl.load_workbook(str(p), data_only=True)
        assert "Data Sheet" in wb.sheetnames

    def test_headers(self, tmp_path):
        p = tmp_path / "PS 2026-06-26.xlsx"
        render_portfolio_solutions(_ps_rows(), p)
        wb = openpyxl.load_workbook(str(p), data_only=True)
        ws = wb["Data Sheet"]
        headers = [ws.cell(1, c).value for c in range(1, 9)]
        assert headers == [
            "Date", " RANK", "TICKER",
            "1-WEEKCHANGE", "1-MONTHCHANGE", "ENTRYDATE",
            "ASSET CLASS", "POSITIONSIZING",
        ]

    def test_rank_col_leading_space(self, tmp_path):
        """' RANK' header (leading space) must be preserved for Excel parity."""
        p = tmp_path / "PS 2026-06-26.xlsx"
        render_portfolio_solutions(_ps_rows(1), p)
        wb = openpyxl.load_workbook(str(p), data_only=True)
        ws = wb["Data Sheet"]
        assert ws.cell(1, 2).value == " RANK"

    def test_row_count(self, tmp_path):
        p = tmp_path / "PS 2026-06-26.xlsx"
        render_portfolio_solutions(_ps_rows(4), p)
        wb = openpyxl.load_workbook(str(p), data_only=True)
        ws = wb["Data Sheet"]
        assert ws.max_row == 5  # header + 4 rows

    def test_rank_value(self, tmp_path):
        p = tmp_path / "PS 2026-06-26.xlsx"
        render_portfolio_solutions(_ps_rows(2), p)
        wb = openpyxl.load_workbook(str(p), data_only=True)
        ws = wb["Data Sheet"]
        assert ws.cell(2, 2).value == 1
        assert ws.cell(3, 2).value == 2

    def test_ticker_value(self, tmp_path):
        p = tmp_path / "PS 2026-06-26.xlsx"
        render_portfolio_solutions(_ps_rows(1), p)
        wb = openpyxl.load_workbook(str(p), data_only=True)
        ws = wb["Data Sheet"]
        assert ws.cell(2, 3).value == "PS0"


# ── render_the_call ───────────────────────────────────────────────────────────

class TestRenderTheCall:
    def test_creates_csv_file(self, tmp_path):
        p = tmp_path / "call 2026-06-26.csv"
        render_the_call(_call_rows(), p)
        assert p.exists()

    def test_headers(self, tmp_path):
        p = tmp_path / "call 2026-06-26.csv"
        render_the_call(_call_rows(), p)
        with open(str(p), newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            headers = next(reader)
        assert headers == ["Date", "Symbol", "Outlook", "Outlook Modifier"]

    def test_row_count(self, tmp_path):
        p = tmp_path / "call 2026-06-26.csv"
        render_the_call(_call_rows(5), p)
        with open(str(p), newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
        assert len(rows) == 6  # header + 5 data rows

    def test_date_format_no_zero_padding(self, tmp_path):
        """Date must be M/D/YYYY (no leading zeros) to match real call CSV."""
        p = tmp_path / "call 2026-06-26.csv"
        render_the_call(_call_rows(1), p)
        with open(str(p), newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)  # skip header
            row = next(reader)
        # date(2026, 6, 26) → "6/26/2026"
        assert row[0] == "6/26/2026"

    def test_symbol_col(self, tmp_path):
        p = tmp_path / "call 2026-06-26.csv"
        render_the_call(_call_rows(1), p)
        with open(str(p), newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)
            row = next(reader)
        assert row[1] == "CALL0"

    def test_outlook_col(self, tmp_path):
        p = tmp_path / "call 2026-06-26.csv"
        render_the_call(_call_rows(2), p)
        with open(str(p), newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)
            rows = list(reader)
        assert rows[0][2] == "long"
        assert rows[1][2] == "short"

    def test_empty_rows_writes_header_only(self, tmp_path):
        p = tmp_path / "call 2026-06-26.csv"
        render_the_call([], p)
        with open(str(p), newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
        assert len(rows) == 1
        assert rows[0] == ["Date", "Symbol", "Outlook", "Outlook Modifier"]

    def test_date_first_of_month_no_zero(self, tmp_path):
        """date(2026, 1, 5) → '1/5/2026' — both month and day unpadded."""
        rows = [{"snapshot_date": date(2026, 1, 5), "symbol": "X", "outlook": "long"}]
        p = tmp_path / "call 2026-01-05.csv"
        render_the_call(rows, p)
        with open(str(p), newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)
            row = next(reader)
        assert row[0] == "1/5/2026"


# ── FILE_LANES constant ───────────────────────────────────────────────────────

def test_file_lanes_contains_all_5():
    from etl.hedgeye.emit import FILE_LANES
    expected = {
        ("risk_range", "hist_rr"),
        ("investing_ideas", "hist_iichg"),
        ("etf_changes", "hist_etfchg"),
        ("portfolio_solutions", "hist_ps"),
        ("the_call", "hist_call"),
    }
    assert FILE_LANES == expected


def test_hist_call_top5_not_in_file_lanes():
    """hist_call_top5 is email-only; must stay on direct insert."""
    from etl.hedgeye.emit import FILE_LANES
    assert ("the_call", "hist_call_top5") not in FILE_LANES


if __name__ == "__main__":
    raise SystemExit(pytest.main([__file__, "-v"]))
