"""
Tests for TASK_102 — emit.py race-condition fix and end-to-end backfill proof.

Covers:
1. Code-structure: _register_origin takes NO session parameter, opens own session scope.
2. Code-structure: write_feed calls _register_origin BEFORE writing the file to disk.
3. Integration: 3 emitted Archive files exist with correct sizes (backfill proof).
4. Integration: IIChange 2026-06-24.xlsx has correct sheet/headers/data (format proof).
5. DB: meta_file_origin has 3 email-sourced entries (race-fix proof).
6. DB: source_kind='email' for the 3 generated files (race-fix proof).
7. DB: hist row counts match DEV_HANDOFF expected values.
8. DB: ledger repopulated for all 5 feed types (status=ok).
"""
from __future__ import annotations

import inspect
import os
import sys
from pathlib import Path

import pytest

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).parent.parent

ARCHIVE_FILES = {
    "IIChange": Path(r"C:\Ashok\Investing\Stocks\IIChange\Archive\IIChange 2026-06-24.xlsx"),
    "ETFChange": Path(r"C:\Ashok\Investing\Stocks\ETFChange\Archive\ETFChange 2026-06-24.xlsx"),
    "call": Path(r"C:\Ashok\Investing\Stocks\Call\Archive\call 2026-06-24.csv"),
}
EXPECTED_SIZES = {
    "IIChange": 4981,
    "ETFChange": 5043,
    "call": 433,
}
BACKUP_DIRS = {
    "RR": Path(r"C:\Ashok\Investing\Stocks\RR\Archive\_backup_test_102"),
    "ETFChange": Path(r"C:\Ashok\Investing\Stocks\ETFChange\Archive\_backup_test_102"),
    "Call": Path(r"C:\Ashok\Investing\Stocks\Call\Archive\_backup_test_102"),
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _db_session():
    """Return session_scope or skip if DB unavailable."""
    env_path = PROJECT_ROOT / ".env"
    if env_path.exists():
        try:
            from dotenv import load_dotenv
            load_dotenv(str(env_path))
        except ImportError:
            pass
    try:
        from etl.db import session_scope
        return session_scope
    except Exception as exc:
        pytest.skip(f"Cannot import etl.db: {exc}")


def _pg_available() -> bool:
    """Return True if Postgres is reachable."""
    try:
        from dotenv import load_dotenv
        load_dotenv(str(PROJECT_ROOT / ".env"))
    except ImportError:
        pass
    pw = os.environ.get("PG_PASSWORD", "")
    try:
        import psycopg
        conn = psycopg.connect(
            f"host=localhost port=5432 dbname=trading user=postgres password={pw}",
            connect_timeout=5,
        )
        conn.close()
        return True
    except Exception:
        return False


# ---------------------------------------------------------------------------
# Part 1 — Code-structure tests (pure Python, no DB, no disk)
# ---------------------------------------------------------------------------

class TestRegisterOriginSignature:
    """_register_origin must NOT accept a session parameter (race-fix spec)."""

    def test_no_session_param(self):
        """_register_origin should have exactly one parameter: file_path."""
        from etl.hedgeye.emit import _register_origin
        sig = inspect.signature(_register_origin)
        params = list(sig.parameters.keys())
        assert "session" not in params, (
            f"_register_origin must not accept 'session'; got params: {params}"
        )

    def test_takes_file_path_param(self):
        """_register_origin first (and only) parameter should accept a Path."""
        from etl.hedgeye.emit import _register_origin
        sig = inspect.signature(_register_origin)
        params = list(sig.parameters.keys())
        assert len(params) == 1, (
            f"_register_origin should have 1 param (file_path); got {params}"
        )
        assert params[0] == "file_path"

    def test_opens_own_session_internally(self):
        """_register_origin source must reference session_scope (own session)."""
        from etl.hedgeye import emit as emit_mod
        src = inspect.getsource(emit_mod._register_origin)
        assert "session_scope" in src or "_ss" in src, (
            "_register_origin must open its own session via session_scope; not found in source"
        )

    def test_no_session_arg_in_module_calls(self):
        """write_feed must call _register_origin(dest) — not _register_origin(session, dest)."""
        from etl.hedgeye import emit as emit_mod
        src = inspect.getsource(emit_mod.write_feed)
        # Should not pass 'session' as first arg to _register_origin
        assert "_register_origin(session," not in src, (
            "write_feed must call _register_origin(dest), not _register_origin(session, dest)"
        )
        assert "_register_origin(dest)" in src, (
            "write_feed must call _register_origin(dest) — not found in write_feed source"
        )


class TestWriteFeedCallOrder:
    """write_feed must register origin BEFORE writing the file to disk."""

    def test_register_origin_before_renderer(self):
        """In write_feed source, _register_origin call must appear before renderer call."""
        from etl.hedgeye import emit as emit_mod
        src = inspect.getsource(emit_mod.write_feed)
        idx_register = src.find("_register_origin")
        idx_renderer = src.find("renderer(rows, dest)")
        assert idx_register != -1, "_register_origin call not found in write_feed"
        assert idx_renderer != -1, "renderer(rows, dest) call not found in write_feed"
        assert idx_register < idx_renderer, (
            f"_register_origin (pos {idx_register}) must appear BEFORE "
            f"renderer(rows, dest) (pos {idx_renderer}) in write_feed"
        )


# ---------------------------------------------------------------------------
# Part 2 — Archive file existence (integration, no DB)
# ---------------------------------------------------------------------------

class TestArchiveFilesExist:
    """The 3 emitted files must exist on disk with sizes matching the backfill log."""

    @pytest.mark.parametrize("feed_name", ["IIChange", "ETFChange", "call"])
    def test_file_exists(self, feed_name):
        path = ARCHIVE_FILES[feed_name]
        assert path.exists(), f"MISSING: {path}"

    @pytest.mark.parametrize("feed_name", ["IIChange", "ETFChange", "call"])
    def test_file_nonempty(self, feed_name):
        path = ARCHIVE_FILES[feed_name]
        if not path.exists():
            pytest.skip(f"File missing: {path}")
        size = path.stat().st_size
        assert size > 0, f"File is empty: {path}"

    @pytest.mark.parametrize("feed_name", ["IIChange", "ETFChange", "call"])
    def test_file_size_matches_handoff(self, feed_name):
        """File size must match the size recorded in DEV_HANDOFF."""
        path = ARCHIVE_FILES[feed_name]
        if not path.exists():
            pytest.skip(f"File missing: {path}")
        expected = EXPECTED_SIZES[feed_name]
        actual = path.stat().st_size
        assert actual == expected, (
            f"{feed_name}: expected {expected} bytes, got {actual} bytes"
        )


class TestBackupFoldersExist:
    """Backup folders must exist for RR, ETFChange, Call (proves safe rollback)."""

    @pytest.mark.parametrize("feed_name", ["RR", "ETFChange", "Call"])
    def test_backup_dir_exists(self, feed_name):
        d = BACKUP_DIRS[feed_name]
        assert d.is_dir(), f"MISSING backup dir: {d}"

    def test_rr_backup_has_three_files(self):
        d = BACKUP_DIRS["RR"]
        if not d.is_dir():
            pytest.skip("RR backup dir missing")
        files = [f for f in d.iterdir() if f.is_file()]
        assert len(files) == 3, f"RR backup: expected 3 files, got {[f.name for f in files]}"

    def test_etfchange_backup_has_three_files(self):
        d = BACKUP_DIRS["ETFChange"]
        if not d.is_dir():
            pytest.skip("ETFChange backup dir missing")
        files = [f for f in d.iterdir() if f.is_file()]
        assert len(files) == 3, f"ETFChange backup: expected 3 files, got {[f.name for f in files]}"

    def test_call_backup_has_three_files(self):
        d = BACKUP_DIRS["Call"]
        if not d.is_dir():
            pytest.skip("Call backup dir missing")
        files = [f for f in d.iterdir() if f.is_file()]
        assert len(files) == 3, f"Call backup: expected 3 files, got {[f.name for f in files]}"


# ---------------------------------------------------------------------------
# Part 3 — Emitted file format (spot-check IIChange)
# ---------------------------------------------------------------------------

class TestIIChangeFormat:
    """IIChange 2026-06-24.xlsx must have correct sheet/headers/data rows."""

    @pytest.fixture(autouse=True)
    def _require_file(self):
        if not ARCHIVE_FILES["IIChange"].exists():
            pytest.skip(f"IIChange archive file missing: {ARCHIVE_FILES['IIChange']}")

    def test_sheet_name(self):
        import openpyxl
        wb = openpyxl.load_workbook(str(ARCHIVE_FILES["IIChange"]))
        assert "Data Sheet" in wb.sheetnames

    def test_headers(self):
        import openpyxl
        wb = openpyxl.load_workbook(str(ARCHIVE_FILES["IIChange"]))
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, 6)]
        assert headers == ["Date", " Description", " Ticker", " Outlook", " Action"]

    def test_has_at_least_one_data_row(self):
        import openpyxl
        wb = openpyxl.load_workbook(str(ARCHIVE_FILES["IIChange"]))
        ws = wb.active
        assert ws.max_row >= 2, "Expected at least 1 data row (header + 1)"

    def test_ticker_col_populated(self):
        """Ticker (col 3) in the first data row must be non-empty (MDB known from backfill)."""
        import openpyxl
        wb = openpyxl.load_workbook(str(ARCHIVE_FILES["IIChange"]))
        ws = wb.active
        ticker = ws.cell(2, 3).value
        assert ticker is not None and str(ticker).strip() != "", (
            f"Ticker in row 2 col 3 should be non-empty; got {ticker!r}"
        )

    def test_action_col_has_known_value(self):
        """Action (col 5) in the first data row should be 'remove' (from backfill log)."""
        import openpyxl
        wb = openpyxl.load_workbook(str(ARCHIVE_FILES["IIChange"]))
        ws = wb.active
        action = ws.cell(2, 5).value
        assert action in ("add", "remove", "Add", "Remove"), (
            f"Action in row 2 col 5 expected add/remove; got {action!r}"
        )


# ---------------------------------------------------------------------------
# Part 4 — DB state verification (skipped if Postgres unavailable)
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def db_available():
    if not _pg_available():
        pytest.skip("Postgres not reachable — DB tests skipped")


class TestSourceKindEmail:
    """source_kind='email' must be stamped on the 3 emit-generated files."""

    @pytest.fixture(autouse=True)
    def _skip_no_db(self, db_available):
        pass

    def _count_email(self, file_type_upper: str) -> int:
        from sqlalchemy import text
        ss = _db_session()
        with ss() as s:
            return s.execute(
                text(
                    "SELECT COUNT(*) FROM meta_file_processed "
                    "WHERE UPPER(file_type)=:ft AND source_kind='email'"
                ),
                {"ft": file_type_upper},
            ).scalar()

    def test_iichange_source_kind_email(self):
        count = self._count_email("IICHANGE")
        assert count >= 1, f"Expected >=1 IIChange rows with source_kind='email'; got {count}"

    def test_etfchange_source_kind_email(self):
        count = self._count_email("ETFCHANGE")
        assert count >= 1, f"Expected >=1 ETFChange rows with source_kind='email'; got {count}"

    def test_call_source_kind_email(self):
        count = self._count_email("CALL")
        assert count >= 1, f"Expected >=1 call rows with source_kind='email'; got {count}"


class TestMetaFileOrigin:
    """meta_file_origin must have exactly 3 rows with source_kind='email'."""

    @pytest.fixture(autouse=True)
    def _skip_no_db(self, db_available):
        pass

    def test_has_three_email_entries(self):
        from sqlalchemy import text
        ss = _db_session()
        with ss() as s:
            count = s.execute(
                text("SELECT COUNT(*) FROM meta_file_origin WHERE source_kind='email'")
            ).scalar()
        assert count == 3, (
            f"meta_file_origin: expected 3 email entries, got {count}"
        )

    def test_iichange_origin_registered(self):
        from sqlalchemy import text
        ss = _db_session()
        with ss() as s:
            count = s.execute(
                text(
                    "SELECT COUNT(*) FROM meta_file_origin "
                    "WHERE source_kind='email' AND file_path ILIKE '%IIChange%'"
                )
            ).scalar()
        assert count >= 1, "IIChange file not found in meta_file_origin with source_kind='email'"

    def test_etfchange_origin_registered(self):
        from sqlalchemy import text
        ss = _db_session()
        with ss() as s:
            count = s.execute(
                text(
                    "SELECT COUNT(*) FROM meta_file_origin "
                    "WHERE source_kind='email' AND file_path ILIKE '%ETFChange%'"
                )
            ).scalar()
        assert count >= 1, "ETFChange file not found in meta_file_origin with source_kind='email'"

    def test_call_origin_registered(self):
        from sqlalchemy import text
        ss = _db_session()
        with ss() as s:
            count = s.execute(
                text(
                    "SELECT COUNT(*) FROM meta_file_origin "
                    "WHERE source_kind='email' AND file_path ILIKE '%call%'"
                )
            ).scalar()
        assert count >= 1, "call file not found in meta_file_origin with source_kind='email'"


class TestHistRowCounts:
    """hist table row counts for the test window must match DEV_HANDOFF values."""

    @pytest.fixture(autouse=True)
    def _skip_no_db(self, db_available):
        pass

    def _count(self, table: str, date_col: str) -> int:
        from sqlalchemy import text
        ss = _db_session()
        with ss() as s:
            return s.execute(
                text(
                    f"SELECT COUNT(*) FROM {table} "
                    f"WHERE {date_col} BETWEEN '2026-06-24' AND '2026-06-26'"
                )
            ).scalar()

    def test_hist_iichg_three_rows(self):
        count = self._count("hist_iichg", "event_date")
        assert count == 3, f"hist_iichg: expected 3 rows, got {count}"

    def test_hist_etfchg_six_rows(self):
        count = self._count("hist_etfchg", "event_date")
        assert count == 6, f"hist_etfchg: expected 6 rows, got {count}"

    def test_hist_call_forty_rows(self):
        count = self._count("hist_call", "snapshot_date")
        assert count == 40, f"hist_call: expected 40 rows, got {count}"


class TestLedgerRepopulated:
    """meta_hedgeye_msg must have status=ok entries for all 5 feed types."""

    @pytest.fixture(autouse=True)
    def _skip_no_db(self, db_available):
        pass

    @pytest.mark.parametrize("email_type", [
        "risk_range", "investing_ideas", "etf_changes",
        "portfolio_solutions", "the_call",
    ])
    def test_ledger_has_ok_entry(self, email_type):
        from sqlalchemy import text
        ss = _db_session()
        with ss() as s:
            count = s.execute(
                text(
                    "SELECT COUNT(*) FROM meta_hedgeye_msg "
                    "WHERE email_type=:et AND status='ok'"
                ),
                {"et": email_type},
            ).scalar()
        assert count >= 1, (
            f"meta_hedgeye_msg: expected >=1 ok entry for {email_type!r}; got {count}"
        )
