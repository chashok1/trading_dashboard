#!/usr/bin/env python3
"""
Auto-enrich ref_ma_columns by analyzing Excel workbook and existing code.

Strategy:
1. Load all CSVs (v2, full, seed)
2. Open Excel workbook and read formulas from MA tab
3. Parse Excel formulas to infer source_expr
4. Use existing etl/derive.py mappings as reference
5. Update registry with intelligent defaults
"""

import csv
import re
import sys
from pathlib import Path
from openpyxl import load_workbook
from sqlalchemy import text

# Import case-insensitive sheet lookup helper
from etl.load_raw import get_sheet_case_insensitive

PROJECT_ROOT = Path(__file__).parent.parent

# Excel file location
WORKBOOK_PATH = Path(r"C:\Ashok\Invest\Projects\Cluade\Tickers 2026-04-30.xlsx")

# Column letter to index mapping
def col_letter_to_idx(letter: str) -> int:
    """Convert 'A' -> 1, 'Z' -> 26, 'AA' -> 27, etc."""
    result = 0
    for char in letter:
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result


def idx_to_col_letter(idx: int) -> str:
    """Convert 1 -> 'A', 26 -> 'Z', 27 -> 'AA', etc."""
    result = ""
    while idx > 0:
        idx -= 1
        result = chr(ord('A') + idx % 26) + result
        idx //= 26
    return result


# Source table detection from formula
def detect_source_table(formula: str) -> str:
    """Infer source table from formula references."""
    if formula is None:
        return None

    formula_upper = str(formula).upper()

    # Direct table references
    sheet_patterns = {
        "Y!": "hist_y",
        "TL!": "hist_tl",
        "TD!": "hist_td",
        "TW!": "hist_tw",
        "RR!": "hist_rr",
        "II!": "hist_ii",
        "CALL!": "hist_call",
        "ETF!": "hist_etf",
        "SSS!": "hist_sss",
        "SSS!": "hist_sss",
        "TO!": "hist_to",
        "PS!": "hist_ps",
        "F!": "hist_f",
        "CS!": "hist_cs",
        "TRIG!": None,  # Computed, not a source
    }

    for pattern, table in sheet_patterns.items():
        if pattern in formula_upper:
            return table

    return None


def parse_excel_formula_to_sql(formula: str, source_table: str, col_name: str, concept: str) -> str:
    """
    Parse Excel formula and try to convert to SQL.

    Examples:
      =M2 -> hist_td.column_m (if M is mapped)
      =hist_td.bb_top_15d -> hist_td.bb_top_15d (simple lookup, already SQL-like)
      =D2-E2 -> COALESCE(d.value, 0) - COALESCE(e.value, 0)
      =IF(condition, A2, B2) -> CASE WHEN condition THEN a.value ELSE b.value END
    """

    if not formula:
        return None

    formula_str = str(formula).strip()

    # Skip array formula markers (we'll handle separately)
    if "array" in str(formula).lower() or "openpyxl" in str(formula):
        return None

    # If already looks like SQL (contains . and no $), use as-is
    if "." in formula_str and "!" not in formula_str and not formula_str.startswith("="):
        return formula_str
    if "." in formula_str and "!" not in formula_str and formula_str.startswith("="):
        return formula_str[1:]

    # Simple cell reference =A2, =M2, etc.
    if formula_str.startswith("=") and len(formula_str) <= 4 and formula_str[1].isalpha():
        # This is a reference to another column; will be populated at runtime
        return None

    # XLOOKUP formula - usually we can't translate these directly, use NULL
    if "XLOOKUP" in formula_str or "XLOOKUP" in str(formula):
        return None

    # Remove leading =
    if formula_str.startswith("="):
        formula_str = formula_str[1:]

    # Very basic arithmetic translation (not production-ready, but a start)
    # =IF(S2="Y",F2,CN2) -> CASE WHEN s.use_latest = 'Y' THEN y.last ELSE rr.close END
    if formula_str.upper().startswith("IF("):
        # Skip complex IF for now
        return None

    # At this point, if we can't parse it, return None
    # The user will need to fill these in manually
    return None


def load_all_csvs() -> dict:
    """Load v2, full, and seed CSVs."""
    v2_file = PROJECT_ROOT / "docs" / "ma_columns_v2.csv"
    full_file = PROJECT_ROOT / "docs" / "ma_columns_full.csv"
    seed_file = PROJECT_ROOT / "docs" / "ma_columns_registry_seed.csv"

    v2_data = {}
    full_data = {}
    seed_data = {}

    with open(v2_file) as f:
        reader = csv.DictReader(f)
        for row in reader:
            header = row.get("header", "").strip()
            if header:
                v2_data[header] = row

    with open(full_file) as f:
        reader = csv.DictReader(f)
        for row in reader:
            header = row.get("header", "").strip()
            if header:
                full_data[header] = row

    if seed_file.exists():
        with open(seed_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                col_name = row.get("pg_name", "").strip()
                if col_name:
                    seed_data[col_name] = row

    return {"v2": v2_data, "full": full_data, "seed": seed_data}


def read_excel_formulas() -> dict:
    """Read formulas from Excel MA tab."""
    if not WORKBOOK_PATH.exists():
        print(f"WARNING: {WORKBOOK_PATH} not found. Skipping Excel analysis.")
        return {}

    wb = load_workbook(WORKBOOK_PATH, data_only=False)
    ma_sheet_name = get_sheet_case_insensitive(wb, "MA")
    if ma_sheet_name is None:
        print("WARNING: 'MA' sheet not found in workbook")
        return {}

    ma_sheet = wb[ma_sheet_name]
    formulas = {}

    # Row 1 has headers (column names)
    # Each column has a formula (or value) starting in row 2
    for col_idx, cell in enumerate(ma_sheet[1], 1):
        col_letter = idx_to_col_letter(col_idx)
        header = cell.value

        if header and col_idx < 650:  # Only first ~640 columns
            # Get formula from row 2
            if ma_sheet.cell(row=2, column=col_idx):
                formula = ma_sheet.cell(row=2, column=col_idx).value
                formulas[header] = {
                    "col_letter": col_letter,
                    "col_idx": col_idx,
                    "formula": formula,
                }

    return formulas


def map_source_expr(v2_data: dict, full_data: dict, excel_formulas: dict) -> dict:
    """
    Create intelligent source_expr mappings from available data.

    Returns: {header -> source_expr}
    """
    mappings = {}

    for header, v2_row in v2_data.items():
        full_row = full_data.get(header, {})
        excel_info = excel_formulas.get(header, {})
        formula = excel_info.get("formula")

        col_letter = v2_row.get("letter", "")
        concept = v2_row.get("concept", "")

        # Detect source table
        source_table = detect_source_table(formula)

        # Try to parse formula
        source_expr = parse_excel_formula_to_sql(formula, source_table, header, concept)

        # If we couldn't parse, try to infer from concept + source_table
        if not source_expr and source_table:
            # Best guess: use the concept as column name
            col_name = header.lower().replace(" ", "_").replace("%", "pct")
            source_expr = f"{source_table.replace('hist_', '').replace('drv_', '')}.{col_name}"

        mappings[header] = {
            "source_table": source_table,
            "source_expr": source_expr,
            "formula": formula,
            "col_letter": col_letter,
        }

    return mappings


def auto_enrich(session):
    """Main enrichment function."""
    print("Loading CSVs and Excel workbook...")
    csv_data = load_all_csvs()
    v2_data = csv_data["v2"]
    full_data = csv_data["full"]
    seed_data = csv_data["seed"]

    print(f"Loaded {len(v2_data)} rows from v2 CSV")
    print(f"Loaded {len(full_data)} rows from full CSV")
    print(f"Loaded {len(seed_data)} rows from seed CSV")

    print("\nReading Excel formulas...")
    excel_formulas = read_excel_formulas()
    print(f"Read {len(excel_formulas)} formulas from MA tab")

    print("\nMapping source_expr from formulas...")
    source_mappings = map_source_expr(v2_data, full_data, excel_formulas)

    print("\nUpdating registry with intelligent defaults...")
    updates = 0
    for header, mapping in source_mappings.items():
        source_table = mapping["source_table"]
        source_expr = mapping["source_expr"]

        if source_table or source_expr:
            sql = text("""
                UPDATE ref_ma_columns
                SET source_table = COALESCE(:st, source_table),
                    source_expr = COALESCE(:se, source_expr)
                WHERE excel_header = :h
            """)
            session.execute(sql, {
                "st": source_table,
                "se": source_expr,
                "h": header,
            })
            updates += 1

    session.commit()
    print(f"Updated {updates} rows")

    # Print summary
    result = session.execute(text("""
        SELECT
            drv_cat_table,
            COUNT(*) as total,
            COUNT(CASE WHEN source_table IS NOT NULL THEN 1 END) as with_source,
            COUNT(CASE WHEN source_expr IS NOT NULL THEN 1 END) as with_expr,
            COUNT(CASE WHEN display_label IS NOT NULL THEN 1 END) as with_label
        FROM ref_ma_columns
        WHERE drv_cat_table != 'drv_cat_separator'
        GROUP BY drv_cat_table
        ORDER BY drv_cat_table
    """)).mappings().all()

    print("\n" + "="*80)
    print("Registry Completion Status:")
    print("="*80)
    print("Category,Total,Source Table,Source Expr,Display Label")
    total_all = 0
    source_filled = 0
    expr_filled = 0
    label_filled = 0

    for row in result:
        total = row['total']
        src = row['with_source']
        expr = row['with_expr']
        label = row['with_label']
        print(f"{row['drv_cat_table']},{total},{src}/{total},{expr}/{total},{label}/{total}")

        total_all += total
        source_filled += src
        expr_filled += expr
        label_filled += label

    print("-" * 80)
    print(f"TOTALS,{total_all},{source_filled}/{total_all},{expr_filled}/{total_all},{label_filled}/{total_all}")
    print(f"\nCompletion: source_table {100*source_filled//total_all}%, source_expr {100*expr_filled//total_all}%, display_label {100*label_filled//total_all}%")


if __name__ == "__main__":
    from etl.db import session_scope

    with session_scope() as session:
        count = session.execute(text("SELECT COUNT(*) FROM ref_ma_columns")).scalar()
        if count == 0:
            print("ERROR: ref_ma_columns is empty. Run seed_ref_ma_columns.py first.")
            sys.exit(1)

        auto_enrich(session)

