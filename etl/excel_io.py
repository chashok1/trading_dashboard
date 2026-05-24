"""
Excel and CSV reading helpers.

We use openpyxl with data_only=True so formula values are returned (not
formula text). For CSV files, we read with csv.DictReader and populate an
openpyxl Workbook dynamically.
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable, Iterator
import warnings
import csv

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
import zipfile
import tempfile
import shutil


def open_workbook(path: str | Path) -> Workbook:
    """Open Excel (.xlsx) or CSV file and return as openpyxl Workbook.

    Raises RuntimeError if Excel file is corrupted and cannot be read.
    """
    path = Path(path)

    # Suppress only openpyxl's noisy "Workbook contains no default style" warnings
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    if path.suffix.lower() == '.csv':
        # Read CSV and convert to Workbook
        return _csv_to_workbook(path)
    else:
        # Try to read Excel file
        try:
            return load_workbook(filename=str(path), data_only=True, read_only=False)
        except KeyError as e:
            # Corrupted Excel file (missing required entries)
            if "xml" in str(e).lower():
                raise RuntimeError(
                    f"File '{path.name}' is corrupted (missing {e}). "
                    f"Re-save it in Excel and try again."
                ) from e
            raise


def _csv_to_workbook(path: Path) -> Workbook:
    """Read CSV file and return as openpyxl Workbook with single sheet.
    Tries multiple encodings if UTF-8 fails (e.g., Latin-1 for YFiles CSVs).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = path.stem  # Use filename (without extension) as sheet name

    # Try encodings in order; use the first one that works
    encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252', 'utf-16']
    f = None
    for encoding in encodings:
        try:
            f = open(path, 'r', encoding=encoding)
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader, start=1):
                for col_idx, value in enumerate(row, start=1):
                    # Keep all CSV values as strings; casters will handle conversion
                    cell_value = value if value else None
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)
            f.close()
            return wb
        except (UnicodeDecodeError, UnicodeError):
            if f:
                f.close()
            continue

    # If all encodings fail, raise the last error
    raise UnicodeDecodeError('unknown', b'', 0, 1, f'Could not decode {path} with any encoding: {encodings}')


def get_headers(sheet: Worksheet) -> list[str]:
    """Return row 1 header values as strings (None -> ''), with whitespace stripped."""
    out = []
    for c in range(1, sheet.max_column + 1):
        v = sheet.cell(row=1, column=c).value
        out.append("" if v is None else str(v).strip())
    return out


def find_header_indices(sheet: Worksheet, wanted: Iterable[str]) -> dict[str, int]:
    """
    Map each wanted header to its 1-based column index.
    For duplicate headers, we keep the FIRST occurrence; later code may
    look up subsequent ones by index manually. The `wanted` arg is kept for
    API compatibility; callers that need to know which were missing should
    diff `wanted` against the returned dict themselves.
    """
    headers = get_headers(sheet)
    out: dict[str, int] = {}
    for idx, h in enumerate(headers, start=1):
        if h and h not in out:
            out[h] = idx
    return out


def iter_rows_as_dict(sheet: Worksheet, start_row: int = 2) -> Iterator[dict]:
    """
    Yield each row as {header: value}. Stops at first row where ALL cells are None.
    """
    headers = get_headers(sheet)
    last_col = sheet.max_column
    for r in range(start_row, sheet.max_row + 1):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, last_col + 1)]
        if all(v is None for v in row_vals):
            # blank row; assume table ended
            break
        yield {headers[i]: row_vals[i] for i in range(last_col)}
