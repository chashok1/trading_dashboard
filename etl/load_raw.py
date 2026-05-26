"""
Step 1 of the two-step ETL: load raw rows from a workbook into ref_* and
hist_* tables. Skip duplicates (PK conflict = no-op).

Each loader function returns (rows_read, rows_inserted, rows_skipped).

Logging is structured: every load gets a meta_etl_run row.
"""
from __future__ import annotations

import hashlib
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl.workbook import Workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from etl.casters import excel_time_to_hhmm, to_date, to_text
from etl.db import get_table, insert_skip_duplicates, insert_upsert, session_scope
from etl.excel_io import find_header_indices, iter_rows_as_dict, open_workbook
from etl.mappings import HIST_MAPS, REF_MAPS

log = logging.getLogger(__name__)


# =============================================================================
# Small standalone helpers (used by etl_load.py / tickers_initial_load.py)
# =============================================================================

def file_hash(path: str | Path) -> str:
    """Return the SHA-256 hex digest of a file's contents.

    Used by the load audit (meta_file_processed.file_hash) so we can detect
    that a file's bytes haven't changed between scheduler runs and skip
    re-processing.
    """
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 16), b""):
            h.update(chunk)
    return h.hexdigest()


def parse_file_date_from_name(file_name: str) -> Optional[str]:
    """Extract an ISO date (YYYY-MM-DD) from a file name like
    'CS 2026-05-18.csv' or 'F_2026_05_18.csv' or 'Tickers 2026-05-18.xlsx'.

    Returns None when no date pattern is found.
    """
    import re as _re
    m = _re.search(r'(\d{4})[-_/](\d{1,2})[-_/](\d{1,2})', file_name)
    if not m:
        return None
    try:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{y:04d}-{mo:02d}-{d:02d}"
    except Exception:
        return None


# =============================================================================
# Utilities
# =============================================================================

def get_sheet_case_insensitive(wb: Workbook, sheet_name: str) -> Optional[str]:
    """
    Find a sheet in the workbook that matches sheet_name (case-insensitive).
    Returns the actual sheet name or None if not found.
    """
    sheet_lower = sheet_name.lower()
    for actual_sheet in wb.sheetnames:
        if actual_sheet.lower() == sheet_lower:
            return actual_sheet
    return None


def _derive_target_table(file_type: str, target_tab: str) -> Optional[str]:
    """
    Map a (file_type, target_tab) pair from LoadFiles.xlsx to the actual
    DB table the loader writes into. Mirrors the backfill rule in
    db/baseline.sql so newly inserted ref_load_files rows are consistent
    with rows already in the DB.
    """
    if not target_tab:
        return None
    tab = target_tab.strip().lower()
    ft = (file_type or "").strip().lower()
    ref_map = {
        "sctr":           "ref_sector",
        "rrt":            "ref_rrt",
        "desc":           "ref_rule_desc",
        "ismh":           "ref_ismh",
        "miss":           "drv_missing_symbols",
        "ref_load_files": "ref_load_files",
        "loadfiles":      "ref_load_files",
    }
    if tab in ref_map:
        return ref_map[tab]
    if ft == "ref_tickers":
        return None
    return f"hist_{tab}"


# =============================================================================
# Run tracking
# =============================================================================

def open_run(session: Session, *, file_path: str, file_type: str, target_tab: str) -> int:
    table = get_table("meta_etl_run")
    result = session.execute(
        table.insert().values(
            file_path=file_path,
            file_type=file_type,
            target_tab=target_tab,
            status="running",
        ).returning(table.c.run_id)
    )
    run_id = result.scalar_one()
    return run_id


def close_run(session: Session, run_id: int, *,
              rows_read: int, rows_inserted: int, rows_skipped: int,
              skip_reasons: Optional[dict] = None,
              status: str = "success", error_msg: Optional[str] = None) -> None:
    table = get_table("meta_etl_run")
    session.execute(
        table.update()
        .where(table.c.run_id == run_id)
        .values(
            finished_at=datetime.now(),
            rows_read=rows_read,
            rows_inserted=rows_inserted,
            rows_skipped=rows_skipped,
            skip_reasons=skip_reasons,
            status=status,
            error_msg=error_msg,
        )
    )


def update_run_progress(session: Session, run_id: int, rows_inserted: int, rows_skipped: int) -> None:
    """Update meta_etl_run with intermediate progress. Called after each batch commit."""
    table = get_table("meta_etl_run")
    session.execute(
        table.update()
        .where(table.c.run_id == run_id)
        .values(rows_inserted=rows_inserted, rows_skipped=rows_skipped)
    )


# =============================================================================
# Generic mapping-driven loader
# =============================================================================

def _get_raw_value(raw_row: dict, primary_col: str, alternatives: list[str] | None = None) -> str | None:
    """Try to get a value from raw_row, trying alternatives if primary is not found."""
    val = raw_row.get(primary_col)
    if val is not None:
        return val
    if alternatives:
        for alt_col in alternatives:
            val = raw_row.get(alt_col)
            if val is not None:
                return val
    return None


def _row_to_record(raw_row: dict, mapping: dict, source_file: str) -> tuple[Optional[dict], Optional[str]]:
    """
    Apply the column mapping + casters to one row. Returns (record, skip_reason).
    record is None if the row is skipped; skip_reason explains why.
    """
    rec: dict = {}
    cols = mapping["columns"]

    # Apply column mappings (later columns of same db_name override earlier;
    # works for the duplicate-headers cases since the right-most has the value).
    for excel_name, db_name, caster in cols:
        v = raw_row.get(excel_name)
        if caster is not None:
            v = caster(v)
        if v is not None:
            rec[db_name] = v

    # Synthesize snapshot_date / sequence / symbol if needed.
    date_src = mapping.get("date_source_col")
    if date_src and "snapshot_date" not in rec:
        # Try primary column, then CSV alternative "Date"
        date_val = _get_raw_value(raw_row, date_src, ["Date"] if date_src != "Date" else None)
        rec["snapshot_date"] = to_date(date_val)

    seq_src = mapping.get("seq_source_col")
    if seq_src and "sequence" not in rec:
        # Try primary column, then CSV alternative "Time"
        seq_raw = _get_raw_value(raw_row, seq_src, ["Time"] if seq_src != "Time" else None)
        seq_val = excel_time_to_hhmm(seq_raw)
        if seq_val is None:
            seq_val = 0
        rec["sequence"] = seq_val

    sym_src = mapping.get("symbol_source_col")
    if sym_src and "symbol" not in rec:
        sym_val = _get_raw_value(raw_row, sym_src)
        rec["symbol"] = to_text(sym_val)

    # Some hist tabs need post-load fixes (e.g. RR uses y_ticker as symbol)
    aliased = mapping.get("post_load_symbol_from")
    if aliased and not rec.get("symbol"):
        rec["symbol"] = rec.get(aliased)

    # Skip totals/summary rows (e.g., Schwab/Fidelity exports have "Totals" rows).
    # NOTE: "Cash & Cash Investments" used to be in this list but it's an actual
    # cash-balance row (carries the account's cash market_value), not a summary;
    # leaving it in caused the Cash KPI tile to read 0. Keep "Positions Total"
    # and bare "Total/Totals" as summary skips.
    symbol = (rec.get("symbol") or "").strip().lower()
    if symbol in ("total", "totals", "positions total"):
        return None, "summary_row"

    # Validity gate: a row must have whatever the PK requires.
    pk = mapping.get("pk_columns") or []
    for pk_col in pk:
        if pk_col not in rec or rec[pk_col] is None or rec[pk_col] == "":
            return None, f"missing_pk_column_{pk_col}"

    return rec, None


def load_one_tab(session: Session, wb: Workbook, mapping: dict,
                 source_file: str, run_id: Optional[int] = None) -> tuple[int, int, int, dict]:
    """
    Load one Excel tab into its target table. Returns (read, inserted, skipped, skip_reasons_dict).
    Commits per batch and shows progress. If run_id is provided, updates meta_etl_run.
    skip_reasons_dict tracks counts of skipped rows by reason.
    """
    sheet_name = mapping["sheet"]
    table_name = mapping["table"]

    # For CSV files, the sheet name might not match the expected name.
    # Try case-insensitive lookup first. If the expected sheet doesn't exist but there's another sheet, use it.
    actual_sheet_name = get_sheet_case_insensitive(wb, sheet_name)
    if actual_sheet_name is None:
        if len(wb.sheetnames) == 1:
            # CSV file with only one sheet — use it regardless of name.
            # This is the normal happy path for CSVs exported by tools that
            # name the sheet after the source file rather than the canonical
            # tab. Demoted from warning → debug so it doesn't spam every load.
            actual_sheet_name = wb.sheetnames[0]
            log.debug("Sheet '%s' not found; using CSV sheet '%s'", mapping["sheet"], actual_sheet_name)
        else:
            log.warning("Sheet '%s' not in workbook; skipping", sheet_name)
            return 0, 0, 0, {}

    sheet = wb[actual_sheet_name]
    start_row = 2 + mapping.get("skip_first_n", 0)

    add_source_file = table_name.startswith("hist_")
    records: list[dict] = []
    rows_read = 0
    skip_reasons: dict[str, int] = {}

    for raw in iter_rows_as_dict(sheet, start_row=start_row):
        rows_read += 1
        rec, skip_reason = _row_to_record(raw, mapping, source_file)
        if rec is not None:
            if add_source_file:
                rec["source_file"] = source_file
            records.append(rec)
        elif skip_reason:
            skip_reasons[skip_reason] = skip_reasons.get(skip_reason, 0) + 1

    print(f"\n{table_name}: started. tab rows - {rows_read}")

    if not records:
        n_skipped = rows_read
        if skip_reasons:
            print(f"  [{table_name}] All rows skipped. Reasons: {skip_reasons}")
        return rows_read, 0, n_skipped, skip_reasons

    label = f"{sheet_name}->{table_name}"

    def _progress(batch_num: int, total_batches: int, n_ins: int, n_skp: int) -> None:
        if run_id is not None:
            update_run_progress(session, run_id, n_ins, n_skp)
        pct = int(100 * batch_num / total_batches)
        print(f"  [{label}] batch {batch_num}/{total_batches} ({pct}%)  "
              f"cumulative: {n_ins} inserted, {n_skp} skipped")

    if table_name.startswith("hist_"):
        n_attempted, n_inserted = insert_upsert(session, table_name, records, progress_cb=_progress)
    else:
        n_attempted, n_inserted = insert_skip_duplicates(session, table_name, records, progress_cb=_progress)
    n_skipped = n_attempted - n_inserted

    print(f"  [{label}] DONE  read={rows_read}  inserted={n_inserted}  skipped={n_skipped}")
    if skip_reasons:
        print(f"  [{label}] Row skip reasons during parsing: {skip_reasons}")
    log.info("%-20s  %5d read  %5d inserted  %5d skipped",
             label, rows_read, n_inserted, n_skipped)
    return rows_read, n_inserted, n_skipped, skip_reasons


# =============================================================================
# Special-case tab loaders (custom layouts)
# =============================================================================

def load_data_tab_holidays(session: Session, wb: Workbook, source_file: str) -> tuple[int, int, int]:
    """Data!O:P -> ref_holiday."""
    sheet_name = get_sheet_case_insensitive(wb, "Data")
    if sheet_name is None:
        return 0, 0, 0
    sheet = wb[sheet_name]
    total_rows = sheet.max_row - 1
    print(f"\nref_holiday: started. tab rows - {total_rows}")
    rows_read = 0
    records: list[dict] = []
    for r in range(2, sheet.max_row + 1):
        d = to_date(sheet.cell(row=r, column=15).value)   # O
        desc = to_text(sheet.cell(row=r, column=16).value)  # P
        if d is None:
            continue
        rows_read += 1
        records.append({"holiday_date": d, "description": desc})
    n_attempted, n_inserted = insert_skip_duplicates(session, "ref_holiday", records)
    return rows_read, n_inserted, n_attempted - n_inserted


def load_data_tab_econ(session: Session, wb: Workbook, source_file: str) -> tuple[int, int, int]:
    """Data!B:M -> ref_econ_indicator."""
    sheet_name = get_sheet_case_insensitive(wb, "Data")
    if sheet_name is None:
        return 0, 0, 0
    sheet = wb[sheet_name]
    total_rows = sheet.max_row - 1
    print(f"\nref_econ_indicator: started. tab rows - {total_rows}")
    cols = [
        ("url", 2), ("indicator", 3), ("indicator_date", 4), ("days", 5),
        ("ol", 6), ("from_date", 7), ("to_date", 8), ("effective_today", 9),
        ("show_on_dashboard", 10), ("incl", 11), ("show_flag", 12), ("expected", 13),
    ]
    records: list[dict] = []
    rows_read = 0
    for r in range(2, sheet.max_row + 1):
        d = to_date(sheet.cell(row=r, column=4).value)
        ind = to_text(sheet.cell(row=r, column=3).value)
        if not d or not ind:
            continue
        rows_read += 1
        rec = {
            "indicator_date": d,
            "indicator": ind,
            "url": to_text(sheet.cell(row=r, column=2).value),
            "days": _safe_int(sheet.cell(row=r, column=5).value),
            "ol": to_text(sheet.cell(row=r, column=6).value),
            "from_date": to_date(sheet.cell(row=r, column=7).value),
            "to_date": to_date(sheet.cell(row=r, column=8).value),
            "effective_today": _ch1(sheet.cell(row=r, column=9).value),
            "show_on_dashboard": _ch1(sheet.cell(row=r, column=10).value),
            "incl": _ch1(sheet.cell(row=r, column=11).value),
            "show_flag": _ch1(sheet.cell(row=r, column=12).value),
            "expected": to_text(sheet.cell(row=r, column=13).value),
        }
        records.append(rec)
    n_attempted, n_inserted = insert_skip_duplicates(session, "ref_econ_indicator", records)
    return rows_read, n_inserted, n_attempted - n_inserted


def load_data_tab_blackout(session: Session, wb: Workbook, source_file: str) -> tuple[int, int, int]:
    """Data!T:U -> ref_fed_blackout."""
    sheet_name = get_sheet_case_insensitive(wb, "Data")
    if sheet_name is None:
        return 0, 0, 0
    sheet = wb[sheet_name]
    total_rows = sheet.max_row - 1
    print(f"\nref_fed_blackout: started. tab rows - {total_rows}")
    records: list[dict] = []
    rows_read = 0
    for r in range(2, sheet.max_row + 1):
        s = to_date(sheet.cell(row=r, column=20).value)   # T
        e = to_date(sheet.cell(row=r, column=21).value)   # U
        if s is None:
            continue
        rows_read += 1
        records.append({"start_date": s, "end_date": e})
    n_attempted, n_inserted = insert_skip_duplicates(session, "ref_fed_blackout", records)
    return rows_read, n_inserted, n_attempted - n_inserted


# Mapping for Data!R + W..CC each-other-column with header from row 1.
# Skip blank columns. Each header becomes a "category".
_CALENDAR_EVENT_COLS = [18, 23, 25, 27, 29, 31, 33, 35, 37, 39, 41, 43, 45, 47, 49,
                       51, 53, 55, 57, 59, 61, 63, 65, 67, 69, 71, 73, 75, 77, 79, 81]
# R=18, then W=23, Y=25, AA=27, ... CC=81 (every other col from 23-81 plus R)


def load_data_tab_calendar_events(session: Session, wb: Workbook, source_file: str) -> tuple[int, int, int]:
    """Data!R + W..CC every-other-col -> ref_calendar_event."""
    sheet_name = get_sheet_case_insensitive(wb, "Data")
    if sheet_name is None:
        return 0, 0, 0
    sheet = wb[sheet_name]
    total_rows = sheet.max_row - 1
    print(f"\nref_calendar_event: started. tab rows - {total_rows}")
    records: list[dict] = []
    rows_read = 0
    for c in _CALENDAR_EVENT_COLS:
        category = to_text(sheet.cell(row=1, column=c).value)
        if not category:
            continue
        for r in range(2, sheet.max_row + 1):
            d = to_date(sheet.cell(row=r, column=c).value)
            if d is None:
                continue
            rows_read += 1
            records.append({"category": category, "event_date": d})
    n_attempted, n_inserted = insert_skip_duplicates(session, "ref_calendar_event", records)
    return rows_read, n_inserted, n_attempted - n_inserted


def load_loadfiles(session: Session, loadfiles_path: str) -> tuple[int, int, int]:
    """LoadFiles.xlsx Sheet1 -> ref_load_files."""
    wb = open_workbook(loadfiles_path)
    sheet_name = get_sheet_case_insensitive(wb, "Sheet1")
    if sheet_name is None:
        return 0, 0, 0
    sheet = wb[sheet_name]
    total_rows = sheet.max_row - 1
    print(f"\nref_load_files: started. tab rows - {total_rows}")
    records: list[dict] = []
    rows_read = 0
    for r in range(2, sheet.max_row + 1):
        src = to_text(sheet.cell(row=r, column=1).value)
        ftype = to_text(sheet.cell(row=r, column=2).value)
        tab = to_text(sheet.cell(row=r, column=3).value)
        wd = to_text(sheet.cell(row=r, column=4).value)
        ft_v = sheet.cell(row=r, column=5).value
        from etl.casters import to_time as _t
        ft = _t(ft_v)
        if not (src and ftype and tab and wd):
            continue
        rows_read += 1
        records.append({
            "source_dir": src,
            "file_type": ftype,
            "target_tab": tab,
            "week_day": wd,
            "file_time": ft,
            "enabled": True,
            "target_table": _derive_target_table(ftype, tab),
        })
    n_attempted, n_inserted = insert_skip_duplicates(session, "ref_load_files", records)
    return rows_read, n_inserted, n_attempted - n_inserted


# Trig rule definitions (ref_trig_atomic_rule + ref_trig_composite_mapping)
def load_trig_rules(session: Session, wb: Workbook) -> tuple[int, int, int]:
    """
    Trig rows 4-118 -> ref_trig_atomic_rule
    Trig cols O,Q,S,...EM (rows 4-118) -> ref_trig_composite_mapping
    """
    sheet_name = get_sheet_case_insensitive(wb, "Trig")
    if sheet_name is None:
        return 0, 0, 0
    sheet = wb[sheet_name]
    total_rows = min(sheet.max_row, 119) - 3
    print(f"\nref_trig_atomic_rule + ref_trig_composite_mapping: started. tab rows - {total_rows}")

    # 1. Atomic rule defs
    # The Trig tab puts the MA tab column header in col L (col 12). That value
    # is the durable identity of an atomic rule — every workbook revision keeps
    # it. Cols A and B are optional internal labels and are often blank.
    # We accept any row with a non-empty col L; we also mirror col L into
    # rule_name so etl/derive.py's resolver can look it up against
    # ref_ma_columns.excel_header without an extra fallback.
    atom_records: list[dict] = []
    atom_rows_read = 0
    for r in range(4, min(sheet.max_row, 119) + 1):
        ma_col = to_text(sheet.cell(row=r, column=12).value)  # L
        if not ma_col:
            continue
        atom_rows_read += 1
        atom_records.append({
            "atomic_rule_id": r,
            "rule_name":      ma_col,                                       # col L (mirror)
            "brkeout_from":   _safe_num(sheet.cell(row=r, column=3).value),
            "brkeout_to":     _safe_num(sheet.cell(row=r, column=4).value),
            "wt_below":       _safe_num(sheet.cell(row=r, column=5).value),
            "wt_between":     _safe_num(sheet.cell(row=r, column=6).value),
            "wt_above":       _safe_num(sheet.cell(row=r, column=7).value),
            "ma_column_name": ma_col,                                       # col L
        })

    insert_skip_duplicates(session, "ref_trig_atomic_rule", atom_records)
    valid_atomic_ids = {rec["atomic_rule_id"] for rec in atom_records}

    # 2. Composite -> atomic mappings.
    # Composite rule headers live at row 1 in cols O,Q,S,U,W,...
    comp_records: list[dict] = []
    rows_read = 0
    # Scan columns 15 (O) onwards in steps of 2, until we run out
    for c in range(15, sheet.max_column + 1, 2):
        code = to_text(sheet.cell(row=1, column=c).value)
        if not code:
            continue
        # Walk rows 4..max for a non-blank value in this col.
        for r in range(4, min(sheet.max_row, 119) + 1):
            if r not in valid_atomic_ids:
                continue
            mark = sheet.cell(row=r, column=c).value
            wt = sheet.cell(row=r, column=c + 1).value
            if mark is None or (isinstance(mark, str) and mark.strip() == ""):
                continue
            rows_read += 1
            comp_records.append({
                "composite_rule_code": code,
                "atomic_rule_id": r,
                "weight_override": _safe_num(wt),
            })

    n_att, n_ins = insert_skip_duplicates(session, "ref_trig_composite_mapping", comp_records)

    # ─── Pruning pass (2026-05-12) ──────────────────────────────────────────
    # The workbook is the source of truth. Insert-only ON CONFLICT DO NOTHING
    # leaves rows that the user has deleted from Trig forever-active. Mark them
    # deprecated_at = now() so the rule engine ignores them but history is kept.
    workbook_codes = sorted({rec["composite_rule_code"] for rec in comp_records})
    if workbook_codes:
        session.execute(
            text("""
                UPDATE ref_trig_composite_mapping
                   SET deprecated_at = now()
                 WHERE deprecated_at IS NULL
                   AND composite_rule_code != ALL(:codes)
            """),
            {"codes": workbook_codes},
        )

    workbook_atomic_ids = sorted({rec["atomic_rule_id"] for rec in atom_records})
    if workbook_atomic_ids:
        session.execute(
            text("""
                UPDATE ref_trig_atomic_rule
                   SET deprecated_at = now()
                 WHERE deprecated_at IS NULL
                   AND atomic_rule_id != ALL(:ids)
            """),
            {"ids": workbook_atomic_ids},
        )

    return atom_rows_read + rows_read, len(atom_records) + n_ins, max(0, n_att - n_ins)


# =============================================================================
# History loaders for ETF/II "change" tabs.
# These are HISTORY tables (not event logs): per-stock weekly UPDATES to
# entries originally in hist_etf / hist_ii. Append-only, same cadence as
# their parent.
# =============================================================================

def load_etfchg(session: Session, wb: Workbook, source_file: str) -> tuple[int, int, int]:
    """
    etfchg tab -> hist_etfchg.
    History of per-stock weekly UPDATES to entries from hist_etf.
    Supports two formats based on column structure:
    1. Old format: A=Date ... I=Desc J=Symbol K=Outlook L=Change (11+ columns)
    2. New format (ETFChange): Date, Description, Ticker, Outlook, Action (5 columns)
    PK = (event_date, symbol).
    """
    if not wb.sheetnames:
        return 0, 0, 0

    sheet = wb[wb.sheetnames[0]]
    records: list[dict] = []
    rows_read = 0
    total_rows = sheet.max_row - 1
    print(f"\nhist_etfchg: started. tab rows - {total_rows}")

    # Detect format: if column J has data, it's old format; if column 3 has symbol names, it's new format
    is_new_format = sheet.max_column <= 5

    for r in range(2, sheet.max_row + 1):
        if is_new_format:
            # New format: Date (col 1), Description (col 2), Ticker (col 3), Outlook (col 4), Action (col 5)
            d = to_date(sheet.cell(row=r, column=1).value)
            sym = to_text(sheet.cell(row=r, column=3).value)
            desc = to_text(sheet.cell(row=r, column=2).value)
            outlook = to_text(sheet.cell(row=r, column=4).value)
            change = to_text(sheet.cell(row=r, column=5).value)
        else:
            # Old format: A=Date ... I=Desc J=Symbol K=Outlook L=Change
            d = to_date(sheet.cell(row=r, column=1).value)
            sym = to_text(sheet.cell(row=r, column=10).value)
            desc = to_text(sheet.cell(row=r, column=9).value)
            outlook = to_text(sheet.cell(row=r, column=11).value)
            change = to_text(sheet.cell(row=r, column=12).value)

        if d is None or not sym:
            continue
        rows_read += 1
        records.append({
            "event_date":    d,
            "symbol":        sym,
            "description":   desc,
            "outlook":       outlook,
            "change_str":    change,
            "source_file":   source_file,
        })
    n_attempted, n_inserted = insert_upsert(session, "hist_etfchg", records)
    return rows_read, n_inserted, n_attempted - n_inserted


def load_etf(session: Session, wb: Workbook, source_file: str) -> tuple[int, int, int]:
    """
    etf tab -> hist_etf with header-tracked outlook.

    The source file groups ETF holdings by section headers like:
        5/10/2026  BULLISH                                        ← header row
        5/10/2026  Income Short...   BUXX  ...                    ← inherits BULLISH
        5/10/2026  Physical Gold     AAAU  ...                    ← inherits BULLISH
        5/10/2026  BEARISH                                        ← header row
        5/10/2026  Weight Loss Drugs OZEM  ...                    ← inherits BEARISH

    Detection: a header row has Ticker empty AND Sector text starting with
    'BULLISH' or 'BEARISH' (case-insensitive). The header value is captured
    as the current outlook and applied to every following data row until
    the next header.

    Layout (tab columns):
      A=Imported Date  B=Sector  C=Ticker  D=Date Added
      E=Recent Price   F=BRR     G=TRR     H=Asset Class
      I=Include
    """
    sheet_name = get_sheet_case_insensitive(wb, "etf")
    if sheet_name is None:
        # Single-sheet fallback: if file has only one sheet, use it regardless of name
        if len(wb.sheetnames) == 1:
            sheet_name = wb.sheetnames[0]
            log.warning("Sheet 'etf' not found; using single sheet '%s'", sheet_name)
        else:
            return 0, 0, 0
    sheet = wb[sheet_name]
    total_rows = sheet.max_row - 1
    print("\nhist_etf: started. tab rows - " + str(total_rows))

    records: list[dict] = []
    rows_read = 0
    headers_seen = 0
    current_outlook: Optional[str] = None

    for r in range(2, sheet.max_row + 1):
        imp     = to_date(sheet.cell(row=r, column=1).value)
        sector  = to_text(sheet.cell(row=r, column=2).value)
        ticker  = to_text(sheet.cell(row=r, column=3).value)

        # Header row: Ticker empty + Sector contains a label
        if not ticker and sector:
            tag = sector.strip().upper()
            if tag.startswith("BULLISH"):
                current_outlook = "BULLISH"
                headers_seen += 1
                continue
            if tag.startswith("BEARISH"):
                current_outlook = "BEARISH"
                headers_seen += 1
                continue
            # Some other label (e.g. NEUTRAL) — capture in modifier so we don't
            # silently drop it; outlook stays NULL
            current_outlook = None
            headers_seen += 1
            continue

        # Skip blank/junk lines
        if imp is None or not ticker:
            continue

        rows_read += 1
        records.append({
            "snapshot_date":     imp,
            "symbol":            ticker,
            "sector":            sector,
            "date_added":        to_date(sheet.cell(row=r, column=4).value),
            "recent_price":      _safe_num(sheet.cell(row=r, column=5).value),
            "brr":               _safe_num(sheet.cell(row=r, column=6).value),
            "trr":               _safe_num(sheet.cell(row=r, column=7).value),
            "asset_class":       to_text(sheet.cell(row=r, column=8).value),
            "outlook":           current_outlook,
            "source_file":       source_file,
        })

    log.info("hist_etf: %d header rows seen, %d data rows read", headers_seen, rows_read)
    if not records:
        return rows_read, 0, 0
    n_attempted, n_inserted = insert_upsert(session, "hist_etf", records)
    return rows_read, n_inserted, n_attempted - n_inserted


def load_iichg(session: Session, wb: Workbook, source_file: str) -> tuple[int, int, int]:
    """
    IIchg tab -> hist_iichg.
    History of per-stock weekly UPDATES to entries from hist_ii.
    Uses header-based column mapping (supports multiple column name variations).
    PK = (event_date, symbol).  event_date = weekly snapshot of the change.
    """
    sheet_name = get_sheet_case_insensitive(wb, "IIchg")
    if sheet_name is None:
        # Single-sheet fallback: if file has only one sheet, use it regardless of name
        if len(wb.sheetnames) == 1:
            sheet_name = wb.sheetnames[0]
            log.warning("Sheet 'IIchg' not found; using single sheet '%s'", sheet_name)
        else:
            return 0, 0, 0
    sheet = wb[sheet_name]

    total_rows = sheet.max_row - 1
    print(f"\nhist_iichg: started. tab rows - {total_rows}")

    records: list[dict] = []
    rows_read = 0
    for raw in iter_rows_as_dict(sheet, start_row=2):
        # Find Date column (case-insensitive, strip spaces)
        d = None
        for key in raw:
            if key and key.strip().lower() == "date":
                d = to_date(raw[key])
                break

        # Find Symbol/Ticker column (case-insensitive, strip spaces)
        sym = None
        for key in raw:
            if key and key.strip().lower() in ("ticker", "symbol"):
                sym = to_text(raw[key])
                break

        if d is None or not sym:
            continue

        # Find Outlook column
        outlook = None
        for key in raw:
            if key and key.strip().lower() == "outlook":
                outlook = to_text(raw[key])
                break

        # Find Description column
        description = None
        for key in raw:
            if key and key.strip().lower() in ("description", "desc"):
                description = to_text(raw[key])
                break

        # Find Change/Action column
        change_str = None
        for key in raw:
            if key and key.strip().lower() in ("change", "action"):
                change_str = to_text(raw[key])
                break

        rows_read += 1
        records.append({
            "event_date":    d,
            "symbol":        sym,
            "outlook":       outlook,
            "description":   description,
            "change_str":    change_str,
            "source_file":   source_file,
        })

    n_attempted, n_inserted = insert_upsert(session, "hist_iichg", records)
    return rows_read, n_inserted, n_attempted - n_inserted


# =============================================================================
# Sector signal series loaders (sss, ssL)
# =============================================================================





# =============================================================================
# Reference loaders for Parm, HQuad, HQds
# =============================================================================

def load_parm(session: Session, wb: Workbook, source_file: str) -> tuple[int, int, int]:
    """
    Parm tab -> ref_param + ref_param_lookup + ref_asset_allocation.

    Many vertically-stacked sub-tables in this sheet. We split them into:
      ref_param            - simple (sheet, name -> value) pairs
      ref_param_lookup     - multi-column lookups (W Vol Rule, BB Range, etc.)
      ref_asset_allocation - AF-AK (asset class targets)

    PK conflicts are silently skipped (idempotent).
    """
    sheet_name = get_sheet_case_insensitive(wb, "Parm")
    if sheet_name is None:
        return 0, 0, 0
    sheet = wb[sheet_name]
    total_rows = sheet.max_row - 1
    print(f"\nref_param + ref_param_lookup + ref_asset_allocation: started. tab rows - {total_rows}")

    rows_read = 0
    n_inserted = 0
    n_skipped = 0

    # ---------- 1. ref_param: simple (sheet, name -> value) pairs ----------
    param_records: list[dict] = []

    # Field-index maps: TW/TO/TD/TL/YH each have (Sheet, Column, Index/Index2)
    field_idx_sections = [
        ("TW", 2, 4),    # B=name, D=index
        ("TO", 7, 9),
        ("TD", 12, 14),
        ("TL", 17, 19),
        ("YH", 22, 24),
        ("MS", 27, None),
    ]
    for r in range(2, sheet.max_row + 1):
        for sh_name, name_c, idx_c in field_idx_sections:
            name = to_text(sheet.cell(row=r, column=name_c).value)
            if not name:
                continue
            idx = _safe_num(sheet.cell(row=r, column=idx_c).value) if idx_c else None
            param_records.append({
                "sheet":      sh_name,
                "param_name": name,
                "value":      _num_to_str(idx),
            })
        # AC/AD: generic param pairs
        pname = to_text(sheet.cell(row=r, column=29).value)
        pval  = sheet.cell(row=r, column=30).value
        if pname:
            param_records.append({"sheet": "_param", "param_name": pname,
                                  "value": _val_to_str(pval)})
        # BZ/CA: IV-related thresholds
        pname = to_text(sheet.cell(row=r, column=78).value)
        pval  = sheet.cell(row=r, column=79).value
        if pname:
            param_records.append({"sheet": "_iv", "param_name": pname,
                                  "value": _val_to_str(pval)})
        # BE/BF: outlook -> weight
        out = to_text(sheet.cell(row=r, column=57).value)
        wt  = _safe_num(sheet.cell(row=r, column=58).value)
        if out:
            param_records.append({"sheet": "outlook", "param_name": out,
                                  "value": _num_to_str(wt)})
        # BH/BI: rr_outlook -> weight
        out = to_text(sheet.cell(row=r, column=60).value)
        wt  = _safe_num(sheet.cell(row=r, column=61).value)
        if out:
            param_records.append({"sheet": "outlook_rr", "param_name": out,
                                  "value": _num_to_str(wt)})
        # AU/AV: SigBuySell -> Lookup
        n = to_text(sheet.cell(row=r, column=47).value)
        v = _safe_num(sheet.cell(row=r, column=48).value)
        if n:
            param_records.append({"sheet": "sig_buysell", "param_name": n,
                                  "value": _num_to_str(v)})
        # BB/BC: PVVBuySell -> Lookup
        n = to_text(sheet.cell(row=r, column=54).value)
        v = _safe_num(sheet.cell(row=r, column=55).value)
        if n:
            param_records.append({"sheet": "pvv_buysell", "param_name": n,
                                  "value": _num_to_str(v)})
        # AX/AY: BuySellCombo -> Seq (stash AZ Type as a 2nd row sheet=combo_type)
        n = to_text(sheet.cell(row=r, column=50).value)
        v = _safe_num(sheet.cell(row=r, column=51).value)
        t = to_text(sheet.cell(row=r, column=52).value)
        if n:
            param_records.append({"sheet": "buysell_combo", "param_name": n,
                                  "value": _num_to_str(v)})
            if t:
                param_records.append({"sheet": "buysell_combo_type", "param_name": n,
                                      "value": t})

    rows_read += len(param_records)
    a, i = insert_skip_duplicates(session, "ref_param", param_records)
    n_inserted += i
    n_skipped += (a - i)

    # ---------- 2. ref_param_lookup: multi-column lookups ----------
    lookup_records: list[dict] = []

    # 2a. AM-AS BuySell weight table (one block, rows 2..21)
    for r in range(2, sheet.max_row + 1):
        code = to_text(sheet.cell(row=r, column=39).value)   # AM
        if not code:
            continue
        lookup_records.append({
            "table_name":  "buysell",
            "code":        code,
            "short_name":  to_text(sheet.cell(row=r, column=43).value),  # AQ short desc
            "action":      to_text(sheet.cell(row=r, column=40).value),  # AN BuySell
            "seq":         _safe_num(sheet.cell(row=r, column=44).value),  # AR sort seq
            "description": to_text(sheet.cell(row=r, column=45).value),  # AS amount class
            "extra1":      _num_to_str(_safe_num(sheet.cell(row=r, column=41).value)),  # AO weight
            "extra2":      to_text(sheet.cell(row=r, column=42).value),  # AP buysell
        })

    # 2b. BK-BM Trig Range (range_from, range_to, value)
    for r in range(2, sheet.max_row + 1):
        rf = _safe_num(sheet.cell(row=r, column=63).value)  # BK
        if rf is None:
            continue
        lookup_records.append({
            "table_name":  "trig_range",
            "code":        _num_to_str(rf),
            "short_name":  None,
            "action":      None,
            "seq":         None,
            "description": to_text(sheet.cell(row=r, column=65).value),  # BM
            "extra1":      _num_to_str(_safe_num(sheet.cell(row=r, column=64).value)),  # BL range_to
            "extra2":      None,
        })

    # 2c. BO-BQ vertically-stacked Vol Score sub-tables (Vlm, Price zone, etc.)
    # Walk down BO; whenever it starts a new (non-blank) Category, group by it.
    for r in range(2, sheet.max_row + 1):
        cat   = to_text(sheet.cell(row=r, column=67).value)  # BO
        score = _safe_num(sheet.cell(row=r, column=68).value)  # BP
        desc  = to_text(sheet.cell(row=r, column=69).value)  # BQ
        if not cat or score is None:
            continue
        cat_key = cat.lower().replace(" ", "_")
        lookup_records.append({
            "table_name":  f"vol_score:{cat_key}",
            "code":        _num_to_str(score),
            "short_name":  None,
            "action":      None,
            "seq":         score,
            "description": desc,
            "extra1":      None, "extra2": None,
        })

    # 2d. BS-BX vertically-stacked rule tables. Each block starts with a
    #     non-numeric BS header row, then numeric rows until a blank.
    sub_tables = []   # list of (name, header_row)
    last_header = None
    for r in range(1, sheet.max_row + 1):
        v = sheet.cell(row=r, column=71).value  # BS
        if v is None:
            continue
        # Header rows have non-numeric BS like "W Vol Rule", "Id", "TN TD Rule"
        if isinstance(v, str) and not v.lstrip("-").isdigit():
            sub_tables.append((v.strip(), r))
            last_header = (v.strip(), r)

    # Map header strings to canonical table_name keys
    name_map = {
        "W Vol Rule":             "w_vol_rule",
        "Id":                     "vol_action",
        "TN TD Rule":             "tn_td_rule",
        "Bull Risk Range Rule":   "bull_rr_rule",
        "!Bull Risk Range Rule":  "nbull_rr_rule",
        "BB Ranges":              "bb_range",
    }

    # Build (start_row, end_row, table_name) for each sub-table
    blocks = []
    for i, (hdr_text, r) in enumerate(sub_tables):
        end_r = sub_tables[i + 1][1] - 1 if i + 1 < len(sub_tables) else sheet.max_row
        tbl = name_map.get(hdr_text)
        if tbl is None:
            continue
        blocks.append((r + 1, end_r, tbl))

    for start, end, tbl in blocks:
        for r in range(start, end + 1):
            v = sheet.cell(row=r, column=71).value  # BS
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
            # Code is the BS value (numeric or string)
            code = _val_to_str(v)
            lookup_records.append({
                "table_name":  tbl,
                "code":        code,
                "short_name":  to_text(sheet.cell(row=r, column=72).value),  # BT
                "action":      to_text(sheet.cell(row=r, column=73).value),  # BU
                "seq":         _safe_num(sheet.cell(row=r, column=74).value),  # BV
                "description": to_text(sheet.cell(row=r, column=75).value),  # BW
                "extra1":      to_text(sheet.cell(row=r, column=76).value),  # BX
                "extra2":      None,
            })

    # 2e. CC-CF IV Action mapping
    for r in range(2, sheet.max_row + 1):
        n = sheet.cell(row=r, column=81).value  # CC IV Number
        if n is None:
            continue
        code = _val_to_str(_safe_num(n))
        lookup_records.append({
            "table_name":  "iv_action",
            "code":        code,
            "short_name":  to_text(sheet.cell(row=r, column=83).value),  # CE Code
            "action":      _num_to_str(_safe_num(sheet.cell(row=r, column=82).value)),  # CD Action#
            "seq":         _safe_num(sheet.cell(row=r, column=82).value),
            "description": to_text(sheet.cell(row=r, column=84).value),  # CF Desc
            "extra1":      None, "extra2": None,
        })

    # 2f. CI-CK Scenario Score
    for r in range(2, sheet.max_row + 1):
        s = sheet.cell(row=r, column=87).value  # CI ScenarioScore
        if s is None:
            continue
        code = _val_to_str(_safe_num(s))
        lookup_records.append({
            "table_name":  "scenario_action",
            "code":        code,
            "short_name":  None,
            "action":      _num_to_str(_safe_num(sheet.cell(row=r, column=89).value)),  # CK ActionKey
            "seq":         _safe_num(sheet.cell(row=r, column=87).value),
            "description": to_text(sheet.cell(row=r, column=88).value),  # CJ ActionDesc
            "extra1":      None, "extra2": None,
        })

    # 2g. CM-CN final score -> label
    for r in range(2, sheet.max_row + 1):
        sc = sheet.cell(row=r, column=91).value  # CM Score
        if sc is None:
            continue
        code = _val_to_str(_safe_num(sc))
        lookup_records.append({
            "table_name":  "final_label",
            "code":        code,
            "short_name":  to_text(sheet.cell(row=r, column=92).value),  # CN FinalLabel
            "action":      None,
            "seq":         _safe_num(sheet.cell(row=r, column=91).value),
            "description": None,
            "extra1":      None, "extra2": None,
        })

    rows_read += len(lookup_records)
    a, i = insert_skip_duplicates(session, "ref_param_lookup", lookup_records)
    n_inserted += i
    n_skipped += (a - i)

    # ---------- 3. ref_asset_allocation: AF-AK ----------
    alloc_records: list[dict] = []
    for r in range(2, sheet.max_row + 1):
        cat = to_text(sheet.cell(row=r, column=32).value)  # AF
        if not cat:
            continue
        alloc_records.append({
            "category":   cat,
            "min_pct":    _safe_num(sheet.cell(row=r, column=33).value),
            "max_pct":    _safe_num(sheet.cell(row=r, column=34).value),
            "min_dollar": _safe_num(sheet.cell(row=r, column=35).value),
            "max_dollar": _safe_num(sheet.cell(row=r, column=36).value),
            "units":      _safe_num(sheet.cell(row=r, column=37).value),
        })
    rows_read += len(alloc_records)
    a, i = insert_skip_duplicates(session, "ref_asset_allocation", alloc_records)
    n_inserted += i
    n_skipped += (a - i)

    return rows_read, n_inserted, n_skipped


# Helpers used by load_parm
def _num_to_str(v):
    if v is None:
        return None
    try:
        f = float(v)
        if f == int(f):
            return str(int(f))
        return repr(f)
    except (ValueError, TypeError):
        return str(v)


def _val_to_str(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return _num_to_str(v)
    return str(v).strip() or None


# Miss tab has NO loader. It is purely derived data (missing stock symbols
# from MA). drv_missing_symbols is populated by derive_missing_symbols(),
# which scans hist_* tables vs drv_ma at each as_of_date.


def load_hquad(session: Session, wb: Workbook, source_file: str) -> tuple[int, int, int]:
    """
    HQuad tab -> ref_quad_outlook.
    Cols: A=Category B=Sub-Category C=Ticker D=Eco.Sensitivity
          E-H=Quad 1..4 outlooks
          I=M Q Outlook  J=Monthly Quad Outlook (numeric score)
          K=Q Outlook    L=Quarterly Quad Outlook (numeric score)
    PK = (category, sub_category).
    """
    sheet_name = get_sheet_case_insensitive(wb, "HQuad")
    if sheet_name is None:
        return 0, 0, 0
    sheet = wb[sheet_name]
    total_rows = sheet.max_row - 1
    print(f"\nref_quad_outlook: started. tab rows - {total_rows}")
    records: list[dict] = []
    rows_read = 0
    for r in range(2, sheet.max_row + 1):
        cat = to_text(sheet.cell(row=r, column=1).value)
        sub = to_text(sheet.cell(row=r, column=2).value)
        if not cat or not sub:
            continue
        rows_read += 1
        records.append({
            "category":        cat,
            "sub_category":    sub,
            "ticker":          to_text(sheet.cell(row=r, column=3).value),
            "eco_sensitivity": to_text(sheet.cell(row=r, column=4).value),
            "quad1":           to_text(sheet.cell(row=r, column=5).value),
            "quad2":           to_text(sheet.cell(row=r, column=6).value),
            "quad3":           to_text(sheet.cell(row=r, column=7).value),
            "quad4":           to_text(sheet.cell(row=r, column=8).value),
            "m_outlook":       to_text(sheet.cell(row=r, column=9).value),
            "m_score":         _safe_num(sheet.cell(row=r, column=10).value),
            "q_outlook":       to_text(sheet.cell(row=r, column=11).value),
            "q_score":         _safe_num(sheet.cell(row=r, column=12).value),
        })
    n_attempted, n_inserted = insert_skip_duplicates(session, "ref_quad_outlook", records)
    return rows_read, n_inserted, n_attempted - n_inserted


def load_hqds(session: Session, wb: Workbook, source_file: str) -> tuple[int, int, int]:
    """
    HQds tab -> ref_quad_periods.
    Two sub-sections in the same sheet:
      cols A-G: monthly periods   (A=Month, B=Start, C=End, D=Active, E=Quads, F=Start2, G=End2)
      cols I-O: quarterly periods (I=Quarter, J=Start, K=End, L=Active, M=Quads, N=Start2, O=End2)
    We use B (start_date) + C (end_date) for monthly and J/K for quarterly.
    label = the human label (Month or Quarter text).
    quad  = the Quads string.
    PK = (period_type, start_date).
    """
    sheet_name = get_sheet_case_insensitive(wb, "HQds")
    if sheet_name is None:
        return 0, 0, 0
    sheet = wb[sheet_name]
    total_rows = sheet.max_row - 1
    print(f"\nref_quad_periods: started. tab rows - {total_rows}")
    records: list[dict] = []
    rows_read = 0
    for r in range(2, sheet.max_row + 1):
        # Monthly
        m_label = sheet.cell(row=r, column=1).value
        m_start = to_date(sheet.cell(row=r, column=2).value)
        m_end   = to_date(sheet.cell(row=r, column=3).value)
        m_quad  = to_text(sheet.cell(row=r, column=5).value)
        if m_start is not None:
            rows_read += 1
            records.append({
                "period_type": "monthly",
                "start_date":  m_start,
                "end_date":    m_end,
                "quad":        m_quad,
                "label":       to_text(m_label) if m_label is not None else None,
            })
        # Quarterly
        q_label = sheet.cell(row=r, column=9).value
        q_start = to_date(sheet.cell(row=r, column=10).value)
        q_end   = to_date(sheet.cell(row=r, column=11).value)
        q_quad  = to_text(sheet.cell(row=r, column=13).value)
        if q_start is not None:
            rows_read += 1
            records.append({
                "period_type": "quarterly",
                "start_date":  q_start,
                "end_date":    q_end,
                "quad":        q_quad,
                "label":       to_text(q_label) if q_label is not None else None,
            })
    n_attempted, n_inserted = insert_skip_duplicates(session, "ref_quad_periods", records)
    return rows_read, n_inserted, n_attempted - n_inserted


# =============================================================================
# Schwab Transaction CSV Loader
# =============================================================================

def load_cs_transactions(session: Session, csv_path: str, source_file: str) -> tuple[int, int, int]:
    """
    Schwab transaction CSV -> hist_cst.
    Idempotent via PK conflict (account, trade_date, action, symbol, quantity, price).

    Filename format: Rollover_IRA_XXX892_Transactions_20260516-170734.csv
    CSV columns: Date, Action, Symbol, Description, Quantity, Price, Fees & Comm, Amount

    Account matching: Extract last 3 digits from filename and find matching account in hist_cs.
    E.g., Rollover_IRA_XXX892 -> 892 -> matches Rollover_IRA ...892 in hist_cs
    """
    import csv
    import re

    # Extract account identifier from filename
    stem = Path(source_file).stem  # e.g., Rollover_IRA_XXX892_Transactions_20260516-170734
    account_from_file = stem.split('_Transactions_')[0]  # Rollover_IRA_XXX892
    # Extract last 3 digits to match against hist_cs account names
    account_id = re.search(r'(\d{3})$', account_from_file)
    account_last_3 = account_id.group(1) if account_id else None

    # Find the matching account in hist_cs using the last 3 digits
    if account_last_3:
        account_row = session.execute(text("""
            SELECT DISTINCT account FROM hist_cs
            WHERE account LIKE :pattern
            LIMIT 1
        """), {"pattern": f"%{account_last_3}"}).scalar()
        account = account_row or account_from_file
    else:
        account = account_from_file

    def _parse_dollar(s: str) -> float | None:
        """Parse dollar string like '$123.45', '-$123.45', or empty string."""
        if not s:
            return None
        cleaned = re.sub(r'[\$,]', '', s.strip())
        try:
            return float(cleaned) if cleaned else None
        except ValueError:
            return None

    records: list[dict] = []
    rows_read = 0

    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if not row.get('Date') or not row.get('Action'):
                continue

            trade_date = to_date(row['Date'])
            if trade_date is None:
                continue
            rows_read += 1
            records.append({
                'account':     account,
                'trade_date':  trade_date,
                'action':      to_text(row['Action']),
                'symbol':      to_text(row.get('Symbol')) or '',
                'description': to_text(row.get('Description') or ''),
                'quantity':    _parse_dollar(row.get('Quantity') or ''),
                'price':       _parse_dollar(row.get('Price') or ''),
                'fees':        _parse_dollar(row.get('Fees & Comm') or ''),
                'amount':      _parse_dollar(row.get('Amount') or ''),
                'source_file': source_file,
            })

    n_attempted, n_inserted = insert_upsert(
        session, 'hist_cst', records,
        conflict_columns=['account', 'trade_date', 'action',
                          'symbol', 'quantity', 'price'],
    )
    return rows_read, n_inserted, n_attempted - n_inserted


# =============================================================================
# Fidelity Transaction CSV Loader (Accounts_History.csv)
# =============================================================================

# Order matters: more specific patterns first.  Each entry is
# (substring_in_action_text_uppercased, normalized_kind).
_F_ACTION_PATTERNS = [
    ("YOU BOUGHT",                       "BUY"),
    ("REINVESTMENT",                     "BUY"),       # dividend reinvestment
    ("YOU SOLD",                         "SELL"),
    ("DIVIDEND RECEIVED",                "DIV"),
    ("LONG-TERM CAP GAIN",               "DIV"),
    ("SHORT-TERM CAP GAIN",              "DIV"),
    ("INTEREST",                         "INT"),
    ("PURCHASE INTO CORE ACCOUNT",       "CASH"),      # SPAXX / money market sweeps
    ("REDEMPTION FROM CORE ACCOUNT",     "CASH"),
    ("DEPOSIT",                          "CASH"),
    ("WITHDRAWAL",                       "CASH"),
    ("TRANSFER",                         "CASH"),
    ("CONVERSION",                       "CASH"),
    ("ROLLOVER",                         "CASH"),
    ("FEE",                              "FEE"),
]


def _f_action_kind(action_text: str) -> str:
    """Map Fidelity's free-form Action column to a normalized kind.

    Fidelity bundles many activity types into the same descriptive column
    ('YOU BOUGHT FOO (BAR) (Cash)', 'PURCHASE INTO CORE ACCOUNT...', etc.).
    We scan for the first matching keyword above; everything else falls into
    'OTHER'. The raw text is preserved verbatim in the `action` column for
    audit, so misclassifications cost nothing to investigate.
    """
    if not action_text:
        return "OTHER"
    up = action_text.upper()
    for kw, kind in _F_ACTION_PATTERNS:
        if kw in up:
            return kind
    return "OTHER"


def load_f_transactions(session: Session, csv_path: str, source_file: str) -> tuple[int, int, int]:
    """
    Fidelity Accounts_History.csv or History_for_Account_*.csv -> hist_ft.
    Idempotent via PK conflict (account, trade_date, action, symbol, quantity, price).

    Expected CSV header (after the 2 empty preamble rows):
      Run Date, Account, Account Number, Action, Symbol, Description, Type,
      Price ($), Quantity, Commission ($), Fees ($), Accrued Interest ($),
      Amount ($), Settlement Date

    If Account/Account Number columns are missing, extracts account number from
    filename pattern like "History_for_Account_249118149.csv".

    Returns (rows_read, rows_inserted, rows_skipped_as_duplicates).
    """
    import csv
    import re

    # Extract account number from filename if present (e.g., History_for_Account_249118149.csv)
    filename_account = None
    acct_match = re.search(r'History_for_Account_(\d+)', source_file)
    if acct_match:
        filename_account = acct_match.group(1)

    def _parse_dollar(s) -> float | None:
        if s is None:
            return None
        s = str(s).strip()
        if not s:
            return None
        cleaned = re.sub(r"[\$,]", "", s)
        try:
            return float(cleaned)
        except ValueError:
            return None

    def _parse_date_mdy(s) -> "date | None":
        if not s:
            return None
        s = str(s).strip()
        if not s:
            return None
        # Fidelity uses M/D/YYYY or MM/DD/YYYY; fall back to the generic to_date.
        for fmt in ("%m/%d/%Y", "%m-%d-%Y"):
            try:
                from datetime import datetime
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return to_date(s)

    records: list[dict] = []
    rows_read = 0

    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        # Fidelity puts 2 blank lines before the header. Skip leading blanks
        # until we find a line that starts with "Run Date".
        lines = f.readlines()

    header_idx = next(
        (i for i, ln in enumerate(lines) if ln.strip().startswith("Run Date")),
        None,
    )
    if header_idx is None:
        log.warning("Accounts_History.csv: no 'Run Date' header found in %s", source_file)
        return 0, 0, 0

    reader = csv.DictReader(lines[header_idx:])
    for row in reader:
        if not row.get("Run Date") or not row.get("Action"):
            continue
        rows_read += 1
        action_text = (row.get("Action") or "").strip()
        # NOTE: `symbol` is NOT NULL DEFAULT '' in hist_ft, but
        # 401(k) "Exchange In/Out" rows and core-account sweeps have no ticker
        # in the CSV.  to_text("") returns None, which would violate the
        # constraint — so we coerce empty symbols (only) back to "".
        records.append({
            "account":          to_text(row.get("Account") or filename_account or ""),
            "account_number":   to_text(row.get("Account Number") or filename_account or ""),
            "trade_date":       _parse_date_mdy(row.get("Run Date")),
            "settlement_date":  _parse_date_mdy(row.get("Settlement Date")),
            "action":           to_text(action_text),
            "action_kind":      _f_action_kind(action_text),
            "symbol":           to_text(row.get("Symbol")) or "",
            "description":      to_text(row.get("Description") or ""),
            "type":             to_text(row.get("Type") or ""),
            "price":            _parse_dollar(row.get("Price ($)") or ""),
            "quantity":         _parse_dollar(row.get("Quantity") or ""),
            "commission":       _parse_dollar(row.get("Commission ($)") or ""),
            "fees":             _parse_dollar(row.get("Fees ($)") or ""),
            "accrued_interest": _parse_dollar(row.get("Accrued Interest ($)") or ""),
            "amount":           _parse_dollar(row.get("Amount ($)") or ""),
            "source_file":      source_file,
        })

    # `quantity` and `price` can both be NULL for 401(k) Exchange In/Out,
    # dividends, and cash sweeps. The post-2026-05-18 schema uses a surrogate
    # `id` PK + a UNIQUE NULLS NOT DISTINCT constraint on the natural key, so
    # we explicitly tell insert_upsert to dedup on that natural key.
    n_attempted, n_inserted = insert_upsert(
        session, "hist_ft", records,
        conflict_columns=["account", "trade_date", "action",
                          "symbol", "quantity", "price"],
    )
    return rows_read, n_inserted, n_attempted - n_inserted


def load_cs_positions_csv(session: Session, csv_path: str, source_file: str) -> tuple[int, int, int]:
    """
    Schwab positions CSV -> hist_cs (same mapping as Excel CS sheet).
    Idempotent via PK conflict (snapshot_date, account, symbol).

    CSV columns: Section, Date, Symbol, Description, Qty (Quantity), Price,
                 Price Chng $ (Price Change $), Price Chng % (Price Change %),
                 Mkt Val (Market Value), Day Chng $ (Day Change $),
                 Day Chng % (Day Change %), Cost Basis, Gain $ (Gain/Loss $),
                 Gain % (Gain/Loss %), Reinvest?, Reinvest Capital Gains?, Asset Type
    """
    import csv
    import re

    def _parse_currency(s: str) -> float | None:
        """Parse currency string like '$1,234.56', '($123.45)' (negative), '--', or empty."""
        if not s or s.strip() in ('--', ''):
            return None
        s = s.strip()
        # Check for parentheses format (negative accounting notation)
        is_negative = s.startswith('(') and s.endswith(')')
        # Remove $, commas, parentheses, and whitespace
        cleaned = re.sub(r'[\$,%\s()]', '', s)
        try:
            val = float(cleaned) if cleaned else None
            return -val if (is_negative and val is not None) else val
        except ValueError:
            return None

    def _parse_percent(s: str) -> float | None:
        """Parse percent string like '1.23%', '(1.23)%' (negative), '--', or empty."""
        if not s or s.strip() in ('--', ''):
            return None
        s = s.strip()
        # Check for parentheses format (negative accounting notation)
        is_negative = s.startswith('(') and ')' in s
        # Remove %, parentheses, and whitespace
        cleaned = re.sub(r'[%\s()]', '', s)
        try:
            val = float(cleaned) if cleaned else None
            return -val if (is_negative and val is not None) else val
        except ValueError:
            return None

    records: list[dict] = []
    rows_read = 0
    rows_skipped = 0

    # Extract snapshot date from filename as a fallback (e.g. "CS 2026-05-18.csv" -> 2026-05-18).
    # Schwab leaves the Date cell blank for the "Cash & Cash Investments" balance row,
    # so without this fallback that row would be dropped at the validity gate.
    import os, re as _re
    _fn = os.path.basename(csv_path)
    _m = _re.search(r'(\d{4})[-_/](\d{1,2})[-_/](\d{1,2})', _fn)
    fname_date = None
    if _m:
        try:
            fname_date = to_date(f"{_m.group(1)}-{int(_m.group(2)):02d}-{int(_m.group(3)):02d}")
        except Exception:
            fname_date = None

    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Require Section + Symbol; Date can be blank (Schwab leaves it blank
            # for the Cash & Cash Investments row). We backfill from filename below.
            if not row.get('Section') or not row.get('Symbol'):
                continue

            symbol = to_text(row['Symbol']).strip() if row.get('Symbol') else ''
            # Skip true summary rows. "Cash & Cash Investments" is the cash-balance
            # row (carries market_value), not a summary — keep it.
            if symbol.lower() in ('positions total', 'total', 'totals'):
                rows_skipped += 1
                continue

            # Resolve snapshot_date: row's Date column if present, else filename date.
            snap_date = to_date(row['Date']) if row.get('Date') else None
            if snap_date is None:
                snap_date = fname_date
            if snap_date is None:
                # No date anywhere — can't insert (PK requires snapshot_date).
                rows_skipped += 1
                continue

            rows_read += 1
            records.append({
                'snapshot_date':      snap_date,
                'account':            to_text(row['Section']),
                'symbol':             symbol,
                'description':        to_text(row.get('Description') or ''),
                'qty':                _parse_currency(row.get('Qty (Quantity)') or ''),
                'price':              _parse_currency(row.get('Price') or ''),
                'price_chng_dollar':  _parse_currency(row.get('Price Chng $ (Price Change $)') or ''),
                'price_chng_pct':     _parse_percent(row.get('Price Chng % (Price Change %)') or ''),
                'market_value':       _parse_currency(row.get('Mkt Val (Market Value)') or ''),
                'day_chng_dollar':    _parse_currency(row.get('Day Chng $ (Day Change $)') or ''),
                'day_chng_pct':       _parse_percent(row.get('Day Chng % (Day Change %)') or ''),
                'cost_basis':         _parse_currency(row.get('Cost Basis') or ''),
                'gain_dollar':        _parse_currency(row.get('Gain $ (Gain/Loss $)') or ''),
                'gain_pct':           _parse_percent(row.get('Gain % (Gain/Loss %)') or ''),
                'reinvest':           to_text(row.get('Reinvest?') or ''),
                'reinvest_cap_gains': to_text(row.get('Reinvest Capital Gains?') or ''),
                'security_type':      to_text(row.get('Asset Type') or ''),
            })

    n_attempted, n_inserted = insert_skip_duplicates(session, 'hist_cs', records)
    return rows_read, n_inserted, (n_attempted - n_inserted)


# --- tiny helpers used above only --------------------------------------------

def _safe_int(v):
    from etl.casters import to_int
    return to_int(v)


def _safe_num(v):
    from etl.casters import to_numeric
    return to_numeric(v)


def _ch1(v):
    """Take the first char of a string-ish value (used for 1-letter flags)."""
    s = to_text(v) or ''
    return s[:1] if s else None
