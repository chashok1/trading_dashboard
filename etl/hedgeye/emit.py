"""
etl/hedgeye/emit.py — File renderer for the 5 tab-backed Hedgeye feeds.

For each email parsed by parsers.py, render the rows into the exact file
format that the existing loader already understands, drop it in the watched
source_dir, and let the scheduler → loader → derive flow ingest it.

Precedence: if a real file for that feed+date already exists in source_dir,
skip rendering (return None). Only render when absent.

source_kind tagging: after writing, insert into meta_file_origin so
etl_load.mark_processed() can stamp source_kind='email'.

Pure functions (render_*) take rows + path — no DB, no network. Tested in
tests/test_hedgeye_emit.py.
"""
from __future__ import annotations

import csv
import logging
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl

from sqlalchemy import text

log = logging.getLogger("hedgeye.emit")

# Mapping: email_type → (file_type in ref_load_files, file extension)
EMIT_MAP: dict[str, tuple[str, str]] = {
    "risk_range":          ("RR",         "xlsx"),
    "investing_ideas":     ("IIChange",   "xlsx"),
    "etf_changes":         ("ETFChange",  "xlsx"),
    "portfolio_solutions": ("PS",         "xlsx"),
    "the_call":            ("call",       "csv"),
}

# Tables written via the file/loader path for these 5 feeds.
# hist_call_top5 is email-only and stays on direct insert.
EMIT_TABLES: frozenset[str] = frozenset({
    "hist_rr", "hist_iichg", "hist_etfchg", "hist_ps", "hist_call",
})

# (email_type, table) pairs that go via the file/loader path
FILE_LANES: frozenset[tuple[str, str]] = frozenset({
    ("risk_range",          "hist_rr"),
    ("investing_ideas",     "hist_iichg"),
    ("etf_changes",         "hist_etfchg"),
    ("portfolio_solutions", "hist_ps"),
    ("the_call",            "hist_call"),
})


# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------

def _get_source_dir(session, file_type: str) -> Optional[str]:
    """Read source_dir from ref_load_files for this file_type (case-insensitive)."""
    row = session.execute(
        text(
            "SELECT source_dir FROM ref_load_files "
            "WHERE LOWER(file_type)=LOWER(:ft) AND enabled=TRUE LIMIT 1"
        ),
        {"ft": file_type},
    ).first()
    return row[0] if row else None


def _register_origin(session, file_path: Path) -> None:
    """Insert into meta_file_origin so mark_processed() stamps source_kind='email'."""
    session.execute(
        text(
            "INSERT INTO meta_file_origin (file_path, source_kind) "
            "VALUES (:p, 'email') ON CONFLICT (file_path) DO NOTHING"
        ),
        {"p": str(file_path)},
    )


# ---------------------------------------------------------------------------
# Precedence check
# ---------------------------------------------------------------------------

def _file_exists_for_date(source_dir: str, file_type: str, feed_date: date) -> bool:
    """True if any file for this feed+date already exists in source_dir."""
    d_str = feed_date.strftime("%Y-%m-%d")
    p = Path(source_dir)
    if not p.exists():
        return False
    prefix = file_type.lower()
    for f in p.iterdir():
        stem = f.stem.lower()
        if stem.startswith(prefix) and d_str in stem:
            return True
    return False


def _dest_path(source_dir: str, file_type: str, feed_date: date, ext: str) -> Path:
    return Path(source_dir) / f"{file_type} {feed_date.strftime('%Y-%m-%d')}.{ext}"


# ---------------------------------------------------------------------------
# Pure renderers (no DB, no network) — unit-tested in test_hedgeye_emit.py
# ---------------------------------------------------------------------------

def render_risk_range(rows: list[dict], path: Path) -> None:
    """Write RR rows as xlsx matching the real file format.

    Sheet: Table_Section
    Headers: Index, Description, Outlook, BUY TRADE, SELL TRADE, Prev Close, RR Date
    Columns loaded by load_rr (case-insensitive key lookup):
      INDEX → symbol, "date" in key → market_close, "buy" in key → buy_trade,
      "sell" in key → sell_trade, Description → name, Outlook → outlook.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Table_Section"
    ws.append(
        ["Index", "Description", "Outlook", "BUY TRADE", "SELL TRADE", "Prev Close", "RR Date"]
    )
    for r in rows:
        ws.append([
            r.get("symbol") or r.get("tos_symbol") or "",
            r.get("name") or "",
            r.get("outlook") or "",
            r.get("buy_trade"),
            r.get("sell_trade"),
            r.get("last_price"),
            r.get("market_close") or r.get("snapshot_date"),
        ])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


def render_investing_ideas(rows: list[dict], path: Path) -> None:
    """Write IIChange rows as xlsx matching the ETFChange file format exactly.

    Sheet: Data Sheet (mirrors ETFChange by design decision 2026-06-27)
    Headers: Date, ' Description', ' Ticker', ' Outlook', ' Action'
      (leading spaces on cols 2-5, byte-for-byte structurally identical to ETFChange)
    load_iichg: case-insensitive sheet lookup for 'IIchg' won't match 'Data Sheet',
      but the single-sheet fallback triggers because the emitted file has exactly
      one sheet. Header keys are stripped of leading spaces, so Date/Description/
      Ticker/Outlook/Action all resolve correctly regardless of leading spaces.
    Column mapping from parser:
      snapshot_date→Date, description→' Description', symbol→' Ticker',
      side→' Outlook', action→' Action'.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Sheet"
    ws.append(["Date", " Description", " Ticker", " Outlook", " Action"])
    for r in rows:
        ws.append([
            r.get("snapshot_date"),
            r.get("description") or "",
            r.get("symbol") or r.get("tos_symbol") or "",
            r.get("side") or "",
            r.get("action") or "",
        ])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


def render_etf_changes(rows: list[dict], path: Path) -> None:
    """Write ETFChange rows as xlsx matching the real ETFChange file format.

    Sheet: Data Sheet (single-sheet; load_etfchg uses position-based reading)
    Headers: Date, ' Description', ' Ticker', ' Outlook', ' Action'
      (leading spaces on cols 2-5 preserved for Excel workbook import parity)
    load_etfchg new-format: col1=Date→event_date, col2=Desc→description,
      col3=Ticker→symbol, col4=Outlook→outlook, col5=Action→change_str.
    Column mapping from parser:
      snapshot_date→col1, ""→col2, symbol→col3, side→col4, action→col5.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Sheet"
    ws.append(["Date", " Description", " Ticker", " Outlook", " Action"])
    for r in rows:
        ws.append([
            r.get("snapshot_date"),
            r.get("description") or "",
            r.get("symbol") or r.get("tos_symbol") or "",
            r.get("side") or "",
            r.get("action") or "",
        ])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


def render_portfolio_solutions(rows: list[dict], path: Path) -> None:
    """Write PS rows as xlsx matching the real PS file format.

    Sheet: Data Sheet (single-sheet; HIST_MAPS ps uses single-sheet fallback)
    Headers: Date, ' RANK', TICKER, 1-WEEKCHANGE, 1-MONTHCHANGE,
             ENTRYDATE, ASSET CLASS, POSITIONSIZING
    Note: ' RANK' has leading space (stripped by get_headers → 'RANK' → HIST_MAPS match).
    load_one_tab HIST_MAPS ps: Date→snapshot_date, RANK→rank, TICKER→ticker, etc.
    Parser provides: snapshot_date, rank, ticker; rest are optional/None.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Sheet"
    ws.append([
        "Date", " RANK", "TICKER",
        "1-WEEKCHANGE", "1-MONTHCHANGE", "ENTRYDATE", "ASSET CLASS", "POSITIONSIZING",
    ])
    for r in rows:
        ws.append([
            r.get("snapshot_date"),
            r.get("rank"),
            r.get("ticker") or r.get("tos_symbol") or "",
            r.get("wk_ago"),
            r.get("mn_ago"),
            r.get("date_added"),
            r.get("asset_class") or "",
            r.get("position_sizing") or "",
        ])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


def render_the_call(rows: list[dict], path: Path) -> None:
    """Write hist_call rows as CSV matching the real call file format.

    Headers: Date, Symbol, Outlook, Outlook Modifier
    Date format: M/D/YYYY (no zero-padding) — e.g. 6/18/2026.
    load_one_tab HIST_MAPS call: date_source_col fallback 'Date'→snapshot_date,
      Symbol→symbol, Outlook→outlook, Outlook Modifier→outlook_modifier.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(str(path), "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Date", "Symbol", "Outlook", "Outlook Modifier"])
        for r in rows:
            d = r.get("snapshot_date")
            if isinstance(d, datetime):
                d = d.date()
            if isinstance(d, date):
                date_str = f"{d.month}/{d.day}/{d.year}"
            else:
                date_str = str(d) if d else ""
            writer.writerow([
                date_str,
                r.get("symbol") or r.get("tos_symbol") or "",
                r.get("outlook") or "",
                r.get("outlook_modifier") or "",
            ])


_RENDERERS = {
    "risk_range":          render_risk_range,
    "investing_ideas":     render_investing_ideas,
    "etf_changes":         render_etf_changes,
    "portfolio_solutions": render_portfolio_solutions,
    "the_call":            render_the_call,
}


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def write_feed(
    session,
    email_type: str,
    feed_date: date,
    rows: list[dict],
) -> Optional[str]:
    """Render rows for one email_type into the watched source_dir.

    Precedence: if a real file for feed+date already exists, returns None
    (skipped). Only renders and registers when absent.

    Returns the file path string if rendered, None if skipped or on error.
    """
    if not rows:
        return None
    entry = EMIT_MAP.get(email_type)
    if not entry:
        return None
    file_type, ext = entry

    source_dir = _get_source_dir(session, file_type)
    if not source_dir:
        log.warning(
            "emit: no source_dir for file_type=%r (email_type=%r) — skipped",
            file_type, email_type,
        )
        return None

    if _file_exists_for_date(source_dir, file_type, feed_date):
        log.debug(
            "emit: real file exists for %s %s — precedence check: skipped",
            file_type, feed_date,
        )
        return None

    dest = _dest_path(source_dir, file_type, feed_date, ext)
    renderer = _RENDERERS[email_type]
    try:
        renderer(rows, dest)
    except Exception:
        log.exception("emit: render failed for %s %s", email_type, feed_date)
        return None

    _register_origin(session, dest)
    log.info("emit: wrote %s (%d rows)", dest, len(rows))
    return str(dest)
