from openpyxl import load_workbook
from pathlib import Path

files_to_check = {
    'RR': r'C:\Ashok\Investing\Stocks\RR\Archive\RR 2026-05-14.xlsx',
    'II': r'C:\Ashok\Investing\Stocks\II\Archive\II 2026-05-11.xlsx',
    'ETF': r'C:\Ashok\Investing\Stocks\ETF\Archive\ETF 2026-05-10.xlsx',
    'PS': r'C:\Ashok\Investing\Stocks\PS\Archive\ps 2026-05-11.xlsx',
}

print("=" * 80)
print("FILE CONTENTS CHECK - Why 0 rows inserted?")
print("=" * 80)

for name, file_path in files_to_check.items():
    p = Path(file_path)
    if not p.exists():
        print(f"\n[{name}] FILE NOT FOUND: {file_path}")
        continue
    
    print(f"\n[{name}] {p.name}:")
    try:
        wb = load_workbook(p)
        print(f"  Sheets: {wb.sheetnames}")
        
        # Find the right sheet
        sheet_name = None
        for sn in [name, name.lower(), 'Data Sheet', 'Data']:
            if sn in wb.sheetnames:
                sheet_name = sn
                break
        
        if not sheet_name:
            sheet_name = wb.sheetnames[0]  # Use first sheet
        
        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column
        
        print(f"  Active sheet: {sheet_name}")
        print(f"  Dimensions: {max_row} rows x {max_col} columns")
        
        # Show first 3 rows
        print(f"  First row (headers):")
        headers = []
        for col in range(1, min(6, max_col + 1)):
            val = ws.cell(1, col).value
            headers.append(str(val)[:15] if val else "")
        print(f"    {' | '.join(headers)}")
        
        # Data rows
        data_rows = max_row - 1  # Subtract header row
        print(f"  Data rows: {data_rows} (total rows {max_row})")
        
        if data_rows > 0:
            print(f"  Sample data (row 2):")
            for col in range(1, min(6, max_col + 1)):
                val = ws.cell(2, col).value
                print(f"    Col {col}: {val}")
    except Exception as e:
        print(f"  ERROR: {e}")