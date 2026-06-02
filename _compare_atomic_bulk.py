import sys, warnings
sys.path.insert(0, ".")
warnings.filterwarnings("ignore")
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from sqlalchemy import text
from etl.db import session_scope

COL_MAP = {
    "MACDH Direction":"macdh_direction","MACD Direction":"macd_direction",
    "BB Direction":"bb_direction","BB Threshold":"bb_threshold",
    "BBThresh CO Days":"bbthresh_co_days","BBThresh_CO_Days2":"bbthresh_co_days2",
    "Trade Cross Over":"trade_cross_over","Trade-Rule":"trade_rule","!Trade Rule":"not_trade_rule",
    "Trend Cross Over":"trend_cross_over","Trend-Rule":"trend_rule","!Trend Rule":"not_trend_rule",
    "Trend Trade Dep Rule":"trend_trade_dep_rule","TrTn Relation":"trtn_relation",
    "!TrTn Relation":"not_trtn_relation","Trade Trend SD Rule":"trade_trend_sd_rule",
    "BRR% Rule":"brrpct_rule","BRR% LRR":"brrpct_lrr","BRR% R2":"brrpct_r2",
    "BRR% LRR2":"brrpct_lrr2","BRR% TRR":"brrpct_trr",
    "BRR% Puts":"brrpct_puts","BRR% TRR Puts":"brrpct_trr_puts","BRR% Dir":"brrpct_dir",
    "High TRR":"high_trr","Low LRR":"low_lrr","Trend below TRR":"trend_below_trr",
    "LRR above Trade":"lrr_above_trade","TRR_Idx":"trr_idx","MRR_Idx":"mrr_idx","LRR_Idx":"lrr_idx",
    "HVAbsolute":"hvabsolute","IVAbsolute":"ivabsolute",
    "IVPercentile":"ivpercentile","IVPercentile Puts":"ivpercentile_puts",
    "HVPercentile":"hvpercentile","HVPercentile Puts":"hvpercentile_puts",
    "IVHV":"ivhv","IVHV Puts":"ivhv_puts","IVRule":"ivrule",
    "RSI Rule":"rsi_rule","RSI Top":"rsi_top","RSI Puts":"rsi_puts",
    "3m-Low-Rule":"3m_low_rule","3m-Low-Days Rule":"3m_low_days_rule",
    "3mn-High-Rule":"3mn_high_rule","3mn-High-Days Rule":"3mn_high_days_rule","3m-Long":"3m_long",
    "Perf3mn SD Rule":"perf3mn_sd_rule","Perf2M SD Rule":"perf2m_sd_rule",
    "Perf3WK SD Rule":"perf3wk_sd_rule","Perf2WK SD Rule":"perf2wk_sd_rule",
    "Perf3D SD Rule":"perf3d_sd_rule","Perf1D SD Rule":"perf1d_sd_rule",
    "!Perf1D_sd":"not_perf1d_sd","Perf3D_sd_1off":"perf3d_sd_1off",
    "Perf SD Rule":"perf_sd_rule","!Perf SD Rule":"not_perf_sd_rule","!Perf3D Rule":"not_perf3d_rule",
    "BBHighLow_SD Rule":"bbhighlow_sd_rule","BBHighLow Days Rule":"bbhighlow_days_rule",
    "BBStreak Rule":"bbstreak_rule","BBStreakRule1":"bbstreakrule1","BBStreak Rule2":"bbstreak_rule2",
    "BBStreak Days Rule":"bbstreak_days_rule","BBStreak Days Rule2":"bbstreak_days_rule2",
    "BBStreak Days Rule3":"bbstreak_days_rule3","BBStreak Days Rule4":"bbstreak_days_rule4",
    "BB Bull Rule":"bb_bull_rule","BB Bull Puts":"bb_bull_puts",
    "BBHighDays":"bbhighdays","BBLowDays":"bblowdays",
    "MACD Rule":"macd_rule","MACDH Rule":"macdh_rule",
    "MACD and H Rule":"macd_and_h_rule","MACD_BRR Puts":"macd_brr_puts",
    "MACDH_BRR Puts":"macdh_brr_puts","MACD and H Rule Puts":"macd_and_h_rule_puts",
    "MACDH Days":"macdh_days","MACDH Days2":"macdh_days2",
    "Overbought":"overbought","!Overbought":"not_overbought",
    "3mn Outlook":"3mn_outlook","3mn Outlook Days":"3mn_outlook_days",
    "3wk Outlook":"3wk_outlook","3wk Outlook Days":"3wk_outlook_days",
    "!3wk ol":"not_3wk_ol","!3wk ol days":"not_3wk_ol_days",
    "BULL":"bull","!BULL":"not_bull","PerfOrBull":"perforbull","!PerfOrBull":"not_perforbull",
    "50-DMA-Rule":"50_dma_rule","50-DMA-Crossover":"50_dma_crossover",
    "200-DMA-Rule":"200_dma_rule","200-DMA-Crossover":"200_dma_crossover",
    "52-Wk Low Rule":"52_wk_low_rule","52-Wk High Rule":"52_wk_high_rule",
    "BRRTrade":"brrtrade","TRRTrade":"trrtrade",
    "Up Resistance":"up_resistance","Down Resistance":"down_resistance","Earnings":"earnings",
    "VS Price":"vs_price","VS Volume Spike":"vs_volume_spike",
    "VS Volatility":"vs_volatility","VS Days":"vs_days","VS LT Outlook Rule":"vs_lt_outlook_rule",
    "Current Price SD Rule":"current_price_sd_rule",
    "Current Volume Rule":"current_volume_rule","Current Volatility Rule":"current_volatility_rule",
    "Short Term Oulook (If LT Bullish)":"short_term_oulook_if_lt_bullish",
    "Short Term Oulook (If LT Bearish)":"short_term_oulook_if_lt_bearish",
}

# Fills
F_MISMATCH  = PatternFill("solid", fgColor="FFB3B3")  # red   — value differs
F_ONENULL   = PatternFill("solid", fgColor="FFE699")  # amber — one side null
F_NO_DB     = PatternFill("solid", fgColor="BFBFBF")  # gray  — symbol not in DB
F_HDR       = PatternFill("solid", fgColor="1F4E79")  # dark blue header
F_SUM_COL   = PatternFill("solid", fgColor="E2EFDA")  # light green — summary cols
F_SUM_ROW   = PatternFill("solid", fgColor="FFF2CC")  # light yellow — summary rows
HDR_FONT    = Font(color="FFFFFF", bold=True, size=9)
SUM_FONT    = Font(bold=True, size=9)
NORM_FONT   = Font(size=9)
MATCH_FONT  = Font(size=9)

print("Loading atomic.xlsx ...")
wb_in = openpyxl.load_workbook("atomic.xlsx", data_only=True)
ws_in = wb_in.active
all_in = list(ws_in.iter_rows(values_only=True))
headers = [str(c).strip() if c is not None else "" for c in all_in[0]]
data_rows = all_in[1:]
n_data = len(data_rows)
n_cols = len(headers)
sym_idx = headers.index("Symbol") if "Symbol" in headers else 0
print(f"  {n_data} data rows, {n_cols} columns")

print("Loading DB ...")
with session_scope() as s:
    d = s.execute(text("SELECT MAX(as_of_date) FROM drv_cat_atomic_input")).scalar()
    rows = s.execute(text("SELECT * FROM drv_cat_atomic_input WHERE as_of_date=:d"), {"d":d}).mappings().fetchall()
db = {str(r["tos_symbol"]).strip(): dict(r) for r in rows if r["tos_symbol"]}
print(f"  DB date={d}  symbols={len(db)}")

# Column indices for mapped cols (0-based in headers)
mapped_indices = {i: COL_MAP[h] for i, h in enumerate(headers) if h in COL_MAP}

# Summary accumulators
col_matched   = [0] * n_cols
col_unmatched = [0] * n_cols

print("Building output workbook ...")
wb_out = openpyxl.Workbook()
ws = wb_out.active
ws.title = "Comparison"
ws.freeze_panes = "B2"

# ── Header row ────────────────────────────────────────────────────────────────
for ci, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=ci, value=h)
    c.fill = F_HDR; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center", wrap_text=False)
    ws.column_dimensions[c.column_letter].width = 10

# Summary header cols
mc_col = n_cols + 1
uc_col = n_cols + 2
for ci, lbl in [(mc_col, "Matched"), (uc_col, "Unmatched")]:
    c = ws.cell(row=1, column=ci, value=lbl)
    c.fill = F_HDR; c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center")
    ws.column_dimensions[c.column_letter].width = 11

# ── Data rows ─────────────────────────────────────────────────────────────────
for ri, data_row in enumerate(data_rows, 2):
    sym = str(data_row[sym_idx]).strip() if data_row[sym_idx] is not None else ""
    db_row = db.get(sym)
    row_matched = 0
    row_unmatched = 0

    for ci, val in enumerate(data_row, 1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.font = NORM_FONT
        c.alignment = Alignment(horizontal="right" if ci > 1 else "left")

        hi = ci - 1  # 0-based header index
        if hi not in mapped_indices:
            continue  # Symbol col or unmapped — no highlight

        if db_row is None:
            c.fill = F_NO_DB
            row_unmatched += 1
            col_unmatched[hi] += 1
            continue

        db_val = db_row.get(mapped_indices[hi])
        try:
            xv = float(val) if val is not None else None
        except (TypeError, ValueError):
            xv = None
        try:
            dv = float(db_val) if db_val is not None else None
        except (TypeError, ValueError):
            dv = None

        if xv is None and dv is None:
            row_matched += 1
            col_matched[hi] += 1
        elif xv is None or dv is None:
            c.fill = F_ONENULL
            row_unmatched += 1
            col_unmatched[hi] += 1
        elif abs(xv - dv) < 0.001:
            row_matched += 1
            col_matched[hi] += 1
        else:
            c.fill = F_MISMATCH
            row_unmatched += 1
            col_unmatched[hi] += 1

    # Summary cols for this row
    for ci, val, fill in [(mc_col, row_matched, F_SUM_COL), (uc_col, row_unmatched, F_SUM_COL)]:
        c = ws.cell(row=ri, column=ci, value=val)
        c.fill = fill; c.font = SUM_FONT
        c.alignment = Alignment(horizontal="center")

# ── Summary rows at bottom ────────────────────────────────────────────────────
sum_rows = [
    (n_data + 2, "Matched Count",   col_matched,   F_SUM_ROW),
    (n_data + 3, "Unmatched Count", col_unmatched, F_SUM_ROW),
]
for sr, label, counts, fill in sum_rows:
    c = ws.cell(row=sr, column=1, value=label)
    c.fill = fill; c.font = SUM_FONT
    for ci in range(1, n_cols + 1):
        hi = ci - 1
        cv = ws.cell(row=sr, column=ci)
        cv.fill = fill; cv.font = SUM_FONT
        cv.alignment = Alignment(horizontal="center")
        if hi in mapped_indices:
            cv.value = counts[hi]
    # blank summary-col intersection
    for ci in [mc_col, uc_col]:
        cv = ws.cell(row=sr, column=ci)
        cv.fill = fill

# Fix Symbol col width
ws.column_dimensions["A"].width = 18
# Freeze row 1 and col A
ws.freeze_panes = "B2"

out_path = "atomic_comparison.xlsx"
wb_out.save(out_path)
print(f"Saved {out_path}  ({n_data} symbols, {n_cols} columns)")
print("Legend: RED=value mismatch  AMBER=one side null  GRAY=symbol not in DB")
