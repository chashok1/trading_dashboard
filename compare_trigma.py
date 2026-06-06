#!/usr/bin/env python3
"""
compare_trigma.py
Compare TrigMA.xlsx values against the database (drv_cat_atomic_input + drv_trig)
for date 2026-06-04.
"""

import sys
import io
sys.path.insert(0, 'C:/Ashok/Invest/Projects/trading-dashboard')

import openpyxl
import pandas as pd
from decimal import Decimal
from sqlalchemy import create_engine, text
from config.settings import settings

# ── DB columns (in order, matching Excel MA cols 267..379) ─────────────────
# These are the ~113 rule columns from drv_cat_atomic_input.
# Metadata cols excluded: as_of_date, source_run_id, computed_at, tos_symbol, ac
DB_RULE_COLS = [
    'macdh_direction', 'macd_direction', 'bb_direction', 'bb_threshold',
    'bbthresh_co_days', 'bbthresh_co_days2', 'trade_cross_over', 'trade_rule',
    'not_trade_rule', 'trend_cross_over', 'trend_rule', 'not_trend_rule',
    'trend_trade_dep_rule', 'trtn_relation', 'not_trtn_relation',
    'trade_trend_sd_rule', 'brrpct_rule', 'brrpct_lrr', 'brrpct_r2',
    'brrpct_lrr2', 'brrpct_trr', 'brrpct_puts', 'brrpct_trr_puts',
    'brrpct_dir', 'high_trr', 'low_lrr', 'trend_below_trr', 'lrr_above_trade',
    'trr_idx', 'mrr_idx', 'lrr_idx', 'hvabsolute', 'ivabsolute',
    'ivpercentile', 'ivpercentile_puts', 'hvpercentile', 'hvpercentile_puts',
    'ivhv', 'ivhv_puts', 'ivrule', 'rsi_rule', 'rsi_top', 'rsi_puts',
    '3m_low_rule', '3m_low_days_rule', '3mn_high_rule', '3mn_high_days_rule',
    '3m_long', 'perf3mn_sd_rule', 'perf2m_sd_rule', 'perf3wk_sd_rule',
    'perf2wk_sd_rule', 'perf3d_sd_rule', 'perf1d_sd_rule', 'not_perf1d_sd',
    'perf3d_sd_1off', 'perf_sd_rule', 'not_perf_sd_rule', 'not_perf3d_rule',
    'bbhighlow_sd_rule', 'bbhighlow_days_rule', 'bbstreak_rule',
    'bbstreakrule1', 'bbstreak_rule2', 'bbstreak_days_rule',
    'bbstreak_days_rule2', 'bbstreak_days_rule3', 'bbstreak_days_rule4',
    'bb_bull_rule', 'bb_bull_puts', 'bbhighdays', 'bblowdays', 'macd_rule',
    'macdh_rule', 'macd_and_h_rule', 'macd_brr_puts', 'macdh_brr_puts',
    'macd_and_h_rule_puts', 'macdh_days', 'macdh_days2', 'overbought',
    'not_overbought', '3mn_outlook', '3mn_outlook_days', '3wk_outlook',
    '3wk_outlook_days', 'not_3wk_ol', 'not_3wk_ol_days', 'bull', 'not_bull',
    'perforbull', 'not_perforbull', '50_dma_rule', '50_dma_crossover',
    '200_dma_rule', '200_dma_crossover', '52_wk_low_rule', '52_wk_high_rule',
    'brrtrade', 'trrtrade', 'up_resistance', 'down_resistance', 'earnings',
    'vs_price', 'vs_volume_spike', 'vs_volatility', 'vs_days',
    'vs_lt_outlook_rule', 'current_price_sd_rule', 'current_volume_rule',
    'current_volatility_rule', 'short_term_oulook_if_lt_bullish',
    'short_term_oulook_if_lt_bearish',
]
# Excel columns 267..379 (113 columns, matching DB_RULE_COLS by position)
EXCEL_ATOMIC_COLS = list(range(267, 380))  # 267..379 inclusive
assert len(EXCEL_ATOMIC_COLS) == len(DB_RULE_COLS), \
    f"Mismatch: {len(EXCEL_ATOMIC_COLS)} Excel cols vs {len(DB_RULE_COLS)} DB cols"

COMPARE_DATE = None   # None => latest as_of_date in drv_cat_atomic_input (current data)
TOLERANCE = 0.001
IGNORE_COLS = {'earnings'}   # earnings days — ignored per request

# Indexes / futures / macro to exclude. The MA "Symbol" column uses TOS-style
# tickers (NOT the ^-prefixed yahoo ones), so pattern rules + a known bare set.
_NONSTOCK_TICKERS = {
    'SPX', 'IXIC', 'RUT', 'VIX', 'VVIX', 'RVX', 'VXN', 'GVZ', 'OVX', 'MOVE',
    'VXD', 'TYX', 'TNX', 'NYICDX', 'DJI', 'COMP', 'HYG', 'LQD', 'GDAXI', 'N225',
    'DXY',
}


def is_nonstock(sym: str) -> bool:
    """True for indexes / futures / FX / macro (excluded from the compare)."""
    s = (sym or '').strip()
    if not s:
        return True
    if s[:1] in ('$', '^', '/'):          # $DXY, $COMP, ^SPX, /ES
        return True
    if ':' in s or s.endswith('=F') or s.endswith('=X'):  # GDAXI:DE, CL=F, EURUSD=X
        return True
    return s.lstrip('^').upper() in _NONSTOCK_TICKERS      # bare SPX, RUT, ...

def normalize_val(v):
    """Normalize to float/int/None for comparison."""
    if v is None:
        return None
    if isinstance(v, Decimal):
        v = float(v)
    if isinstance(v, float) and v == int(v):
        return int(v)
    return v

def vals_equal(a, b):
    """Return True if values are considered equal."""
    a = normalize_val(a)
    b = normalize_val(b)
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= TOLERANCE
    except (TypeError, ValueError):
        return a == b

def main():
    out = io.StringIO()

    def pr(*args, **kwargs):
        print(*args, **kwargs, file=out)
        print(*args, **kwargs)

    # Resolve the compare date to the CURRENT data (latest as_of_date) unless
    # an explicit COMPARE_DATE was set above.
    global COMPARE_DATE
    if COMPARE_DATE is None:
        _eng = create_engine(settings.sqlalchemy_url)
        with _eng.connect() as _c:
            COMPARE_DATE = str(_c.execute(text(
                "SELECT MAX(as_of_date) FROM drv_cat_atomic_input")).scalar())

    pr("=" * 70)
    pr("TrigMA vs DB Comparison Report")
    pr(f"Date: {COMPARE_DATE}")
    pr("=" * 70)

    # ── Load Excel ─────────────────────────────────────────────────────────
    pr("\nLoading TrigMA.xlsx ...")
    wb = openpyxl.load_workbook('TrigMA.xlsx', data_only=True, read_only=True)
    ws = wb['MA']

    excel_data = {}       # symbol → list of 113 values (order == DB_RULE_COLS)
    excel_composite = {}  # symbol → {rule_code: score}

    # Single streaming pass. read_only mode is fast with iter_rows(values_only)
    # but pathologically slow with random ws.cell() access (~140k lookups), and
    # ws.max_row forces a full pre-scan. iter_rows reads each row once.
    rows_iter = ws.iter_rows(min_row=1, values_only=True)
    header = next(rows_iter)                       # row 1 (headers)
    composite_headers = {}                         # col_idx (1-based) → rule_code
    for col in range(381, 442):
        h = header[col - 1] if col - 1 < len(header) else None
        if h:
            composite_headers[col] = h
    atomic_idx = [c - 1 for c in EXCEL_ATOMIC_COLS]   # 0-based positions
    comp_cols = list(composite_headers.items())

    skipped_symbols = 0
    loaded_symbols = 0
    for row in rows_iter:                          # data rows (row 2+)
        sym = row[1] if len(row) > 1 else None      # col B = Symbol
        if sym is None:
            continue
        sym = str(sym).strip()
        if not sym:
            continue
        # Skip indexes / futures / macro (see is_nonstock()).
        if is_nonstock(sym):
            skipped_symbols += 1
            continue
        excel_data[sym] = [
            (normalize_val(row[i]) if i < len(row) else None) for i in atomic_idx
        ]
        excel_composite[sym] = {
            code: (normalize_val(row[col - 1]) if col - 1 < len(row) else None)
            for col, code in comp_cols
        }
        loaded_symbols += 1

    wb.close()
    pr(f"Excel: {loaded_symbols} tradeable symbols loaded, {skipped_symbols} index/futures skipped")

    # ── Load DB atomic ──────────────────────────────────────────────────────
    pr(f"\nQuerying DB drv_cat_atomic_input for {COMPARE_DATE} ...")
    engine = create_engine(settings.sqlalchemy_url)

    col_list = ', '.join(f'"{c}"' for c in DB_RULE_COLS)
    query = f"""
        SELECT tos_symbol, {col_list}
        FROM drv_cat_atomic_input
        WHERE as_of_date = '{COMPARE_DATE}'
    """
    with engine.connect() as conn:
        result = conn.execute(text(query))
        db_atomic = {}
        for row in result:
            sym = row[0]
            vals = [normalize_val(row[i+1]) for i in range(len(DB_RULE_COLS))]
            db_atomic[sym] = vals

    pr(f"DB atomic: {len(db_atomic)} symbols")

    # ── Load DB composite ───────────────────────────────────────────────────
    pr(f"\nQuerying DB drv_trig for {COMPARE_DATE} ...")
    with engine.connect() as conn:
        result = conn.execute(text(f"""
            SELECT tos_symbol, composite_rule_code, score, triggered
            FROM drv_trig
            WHERE as_of_date = '{COMPARE_DATE}'
        """))
        db_trig = {}  # sym → {rule_code → (score, triggered)}
        for row in result:
            sym, code, score, triggered = row
            if sym not in db_trig:
                db_trig[sym] = {}
            db_trig[sym][code] = (normalize_val(score), triggered)

    pr(f"DB composite: {len(db_trig)} symbols, codes per sample: "
       f"{len(next(iter(db_trig.values())))} ")

    # ── Symbol coverage ─────────────────────────────────────────────────────
    excel_syms = set(excel_data.keys())
    db_atomic_syms = set(db_atomic.keys())
    db_trig_syms = set(db_trig.keys())
    common_atomic = excel_syms & db_atomic_syms
    common_trig = excel_syms & db_trig_syms

    pr(f"\n{'─'*70}")
    pr("SYMBOL COVERAGE")
    pr(f"{'─'*70}")
    pr(f"Excel tradeable symbols:         {len(excel_syms)}")
    pr(f"DB atomic symbols:               {len(db_atomic_syms)}")
    pr(f"DB trig symbols:                 {len(db_trig_syms)}")
    pr(f"Common (atomic compare):         {len(common_atomic)}")
    pr(f"Common (composite compare):      {len(common_trig)}")

    only_excel = excel_syms - db_atomic_syms
    only_db = db_atomic_syms - excel_syms
    if only_excel:
        pr(f"\nIn Excel but NOT in DB ({len(only_excel)}): {sorted(only_excel)[:30]}")
    if only_db:
        pr(f"\nIn DB but NOT in Excel ({len(only_db)}): {sorted(only_db)[:30]}")

    # ── Phase 1: Atomic column comparison ──────────────────────────────────
    pr(f"\n{'─'*70}")
    pr("PHASE 1: ATOMIC RULE COLUMNS (drv_cat_atomic_input)")
    pr(f"{'─'*70}")
    pr(f"Comparing {len(DB_RULE_COLS)} columns across {len(common_atomic)} symbols")
    pr()

    atomic_mismatches = {}  # col_name → count of mismatches
    atomic_mismatch_details = []  # (sym, col, excel_val, db_val)
    perfect_match_syms = 0
    total_cells_compared = 0
    total_cells_mismatch = 0
    per_symbol_mismatches = {}  # sym → count

    for sym in sorted(common_atomic):
        ex_vals = excel_data[sym]
        db_vals = db_atomic[sym]
        sym_mismatches = 0
        for i, (col, ev, dv) in enumerate(zip(DB_RULE_COLS, ex_vals, db_vals)):
            if col in IGNORE_COLS:        # earnings — ignored per request
                continue
            total_cells_compared += 1
            if not vals_equal(ev, dv):
                total_cells_mismatch += 1
                sym_mismatches += 1
                atomic_mismatches[col] = atomic_mismatches.get(col, 0) + 1
                atomic_mismatch_details.append((sym, col, ev, dv))
        if sym_mismatches == 0:
            perfect_match_syms += 1
        per_symbol_mismatches[sym] = sym_mismatches

    match_rate = 100.0 * (total_cells_compared - total_cells_mismatch) / max(total_cells_compared, 1)
    pr(f"Total cells compared:     {total_cells_compared}")
    pr(f"Total mismatches:         {total_cells_mismatch}")
    pr(f"Match rate:               {match_rate:.2f}%")
    pr(f"Perfect-match symbols:    {perfect_match_syms}/{len(common_atomic)}")

    if atomic_mismatches:
        pr(f"\nMismatches by column (top 30):")
        pr(f"  {'Column':<40} {'Mismatch Count':>15}")
        pr(f"  {'─'*40} {'─'*15}")
        for col, cnt in sorted(atomic_mismatches.items(), key=lambda x: -x[1])[:30]:
            pr(f"  {col:<40} {cnt:>15}")

    # Top mismatching symbols
    top_mismatch_syms = sorted(per_symbol_mismatches.items(), key=lambda x: -x[1])[:20]
    if any(v > 0 for _, v in top_mismatch_syms):
        pr(f"\nTop symbols by mismatch count (top 20):")
        pr(f"  {'Symbol':<20} {'Mismatches':>12}")
        pr(f"  {'─'*20} {'─'*12}")
        for sym, cnt in top_mismatch_syms:
            if cnt > 0:
                pr(f"  {sym:<20} {cnt:>12}")

    # Sample mismatch details (first 50)
    if atomic_mismatch_details:
        pr(f"\nSample mismatches (first 50 of {len(atomic_mismatch_details)}):")
        pr(f"  {'Symbol':<20} {'Column':<35} {'Excel':>10} {'DB':>10}")
        pr(f"  {'─'*20} {'─'*35} {'─'*10} {'─'*10}")
        for sym, col, ev, dv in atomic_mismatch_details[:50]:
            pr(f"  {sym:<20} {col:<35} {str(ev):>10} {str(dv):>10}")

    # ── Phase 2: Composite score comparison ────────────────────────────────
    pr(f"\n{'─'*70}")
    pr("PHASE 2: COMPOSITE RULE SCORES (drv_trig)")
    pr(f"{'─'*70}")

    # Excel codes vs DB codes
    excel_codes = set(composite_headers.values())
    all_db_codes = set()
    for sym_codes in db_trig.values():
        all_db_codes.update(sym_codes.keys())
    common_codes = excel_codes & all_db_codes
    only_excel_codes = excel_codes - all_db_codes
    only_db_codes = all_db_codes - excel_codes

    pr(f"Excel composite codes:    {len(excel_codes)}")
    pr(f"DB composite codes:       {len(all_db_codes)}")
    pr(f"Common codes:             {len(common_codes)}")
    if only_excel_codes:
        pr(f"Only in Excel:            {sorted(only_excel_codes)}")
    if only_db_codes:
        pr(f"Only in DB:               {sorted(only_db_codes)}")

    comp_mismatches = {}  # code → count
    comp_mismatch_details = []  # (sym, code, excel_value, excel_fired, db_fired)
    total_comp_cells = 0
    total_comp_mismatch = 0
    perfect_comp_syms = 0
    per_comp_sym_mismatches = {}

    for sym in sorted(common_trig):
        ex_comp = excel_composite.get(sym, {})
        db_comp = db_trig.get(sym, {})
        sym_m = 0
        for code in sorted(common_codes):
            total_comp_cells += 1
            ex_score = ex_comp.get(code)
            db_entry = db_comp.get(code)
            if db_entry is None:
                db_score = None
                db_triggered = None
            else:
                db_score, db_triggered = db_entry

            # Excel value is a DEFICIT = total member weight - fired weight.
            #   value <  10  => every GATE fired => TRIGGERED
            #                   (0 = full gate match, 1-9 = watch match)
            #   value >= 10  => at least one gate unfired => NOT fired
            ex_n = normalize_val(ex_score)
            excel_fired = (ex_n is not None and ex_n < 10)
            db_fired = bool(db_triggered)
            if excel_fired != db_fired:
                total_comp_mismatch += 1
                sym_m += 1
                comp_mismatches[code] = comp_mismatches.get(code, 0) + 1
                comp_mismatch_details.append((sym, code, ex_score, excel_fired, db_fired))
        if sym_m == 0:
            perfect_comp_syms += 1
        per_comp_sym_mismatches[sym] = sym_m

    comp_match_rate = 100.0 * (total_comp_cells - total_comp_mismatch) / max(total_comp_cells, 1)
    pr(f"\nComparing {len(common_codes)} codes across {len(common_trig)} symbols")
    pr(f"Total cells compared:     {total_comp_cells}")
    pr(f"Total mismatches:         {total_comp_mismatch}")
    pr(f"Match rate:               {comp_match_rate:.2f}%")
    pr(f"Perfect-match symbols:    {perfect_comp_syms}/{len(common_trig)}")

    if comp_mismatches:
        pr(f"\nMismatches by composite code (top 30):")
        pr(f"  {'Code':<45} {'Mismatch Count':>15}")
        pr(f"  {'─'*45} {'─'*15}")
        for code, cnt in sorted(comp_mismatches.items(), key=lambda x: -x[1])[:30]:
            pr(f"  {code:<45} {cnt:>15}")

    top_comp_sym = sorted(per_comp_sym_mismatches.items(), key=lambda x: -x[1])[:20]
    if any(v > 0 for _, v in top_comp_sym):
        pr(f"\nTop symbols by composite mismatch count (top 20):")
        pr(f"  {'Symbol':<20} {'Mismatches':>12}")
        pr(f"  {'─'*20} {'─'*12}")
        for sym, cnt in top_comp_sym:
            if cnt > 0:
                pr(f"  {sym:<20} {cnt:>12}")

    if comp_mismatch_details:
        pr(f"\nSample composite mismatches (first 50 of {len(comp_mismatch_details)}):")
        pr(f"  {'Symbol':<20} {'Code':<45} {'ExVal':>7} {'ExFired':>8} {'DBFired':>8}")
        pr(f"  {'─'*20} {'─'*45} {'─'*7} {'─'*8} {'─'*8}")
        for sym, code, exv, exf, dbf in comp_mismatch_details[:50]:
            pr(f"  {sym:<20} {code:<45} {str(exv):>7} {str(exf):>8} {str(dbf):>8}")

    # ── Summary ─────────────────────────────────────────────────────────────
    pr(f"\n{'='*70}")
    pr("SUMMARY")
    pr(f"{'='*70}")
    pr(f"Phase 1 (atomic) :  {match_rate:.2f}% match  "
       f"({total_cells_mismatch} mismatches across {len(atomic_mismatches)} columns, "
       f"{len(common_atomic)} symbols)")
    pr(f"Phase 2 (composite): {comp_match_rate:.2f}% match  "
       f"({total_comp_mismatch} mismatches across {len(comp_mismatches)} codes, "
       f"{len(common_trig)} symbols)")

    return out.getvalue()

if __name__ == '__main__':
    report = main()
    with open('trigma_report.txt', 'w', encoding='utf-8') as f:
        f.write(report)
    print("\n\nReport written to trigma_report.txt")
