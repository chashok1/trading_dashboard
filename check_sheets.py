from openpyxl import load_workbook
from pathlib import Path

files = [
    r"C:\Ashok\Invest\Cluade\ETF 2026-05-10.xlsx",
    r"C:\Ashok\Invest\Cluade\ETFChange 2026-05-11.xlsx",
    r"C:\Ashok\Invest\Cluade\II 2026-05-11.xlsx",
    r"C:\Ashok\Invest\Cluade\ps 2026-05-11.xlsx",
    r"C:\Ashok\Invest\Cluade\Tickers 2026-05-12.xlsx",
]

for file in files:
    if Path(file).exists():
        try:
            wb = load_workbook(file)
            name = Path(file).name
            print(f"{name:40} -> {wb.sheetnames}")
        except Exception as e:
            print(f"{file:40} -> ERROR: {e}")
