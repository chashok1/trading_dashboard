#!/usr/bin/env python
"""Extract Trig sheet rules from smaller Excel file."""
from openpyxl import load_workbook

wb = load_workbook("C:\\Ashok\\Invest\\Projects\\Cluade\\Tickers 2026-04-30.xlsx", data_only=False)

if "Trig" in wb.sheetnames:
    trig = wb["Trig"]

    print("Row | Rule Name (Col A) | MA Column (Col B) | From (Col D) | To (Col E)\n")

    derived_indicators = []  # Rules with no From/To
    atomic_rules = []  # Rules with From/To

    for row_idx in range(4, 121):
        col_a = trig.cell(row_idx, 1).value  # Rule name
        col_b = trig.cell(row_idx, 2).value  # MA Column
        col_d = trig.cell(row_idx, 4).value  # From
        col_e = trig.cell(row_idx, 5).value  # To

        if col_a and str(col_a).strip():
            rule_name = str(col_a).strip()
            ma_col = str(col_b).strip() if col_b else ""
            from_val = col_d
            to_val = col_e

            # Check if this is derived (no From/To) or atomic (has From/To)
            if from_val is None and to_val is None:
                derived_indicators.append((rule_name, ma_col))
                print(f"{row_idx:3d} | {rule_name:45s} | {ma_col:30s} | [DERIVED]")
            else:
                atomic_rules.append((rule_name, ma_col, from_val, to_val))
                print(f"{row_idx:3d} | {rule_name:45s} | {ma_col:30s} | [{from_val}, {to_val}]")

    print(f"\n\nSummary:")
    print(f"Derived indicators: {len(derived_indicators)}")
    print(f"Atomic rules: {len(atomic_rules)}")

    print(f"\n\nDerived Indicators (to insert into ref_trig_atomic_rule with NULL From/To):")
    for i, (rule_name, ma_col) in enumerate(derived_indicators, 1):
        print(f"{i:2d}. {rule_name:50s} -> {ma_col}")
