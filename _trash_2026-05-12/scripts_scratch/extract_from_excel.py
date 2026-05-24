#!/usr/bin/env python
"""Extract derived indicators from Trig sheet in Excel."""
from openpyxl import load_workbook

wb = load_workbook("C:\\Ashok\\Invest\\Projects\\Cluade\\Tickers 2026-04-30.xlsx", data_only=False)
trig = wb["Trig"]

derived = []
atomic = []

for row_idx in range(4, 200):
    rule_name = trig.cell(row_idx, 1).value
    ma_column = trig.cell(row_idx, 2).value
    from_val = trig.cell(row_idx, 4).value
    to_val = trig.cell(row_idx, 5).value

    if rule_name and str(rule_name).strip():
        rule_name = str(rule_name).strip()
        ma_column = str(ma_column).strip() if ma_column else ""

        if from_val is None and to_val is None:
            derived.append((ma_column, rule_name))
        else:
            atomic.append((ma_column, rule_name, from_val, to_val))

print(f"Derived indicators found: {len(derived)}")
print(f"Atomic rules found: {len(atomic)}\n")

print("DERIVED INDICATORS (to insert):")
for i, (ma_col, rule_name) in enumerate(derived, 1):
    print(f"{i:2d}. ({repr(ma_col)}, {repr(rule_name)}),")

print(f"\n\nTotal derived: {len(derived)}")
