#!/usr/bin/env python
"""Extract formulas from Trig sheet in Excel for derived indicators."""
from openpyxl import load_workbook

wb = load_workbook("C:\\Ashok\\Invest\\Projects\\Cluade\\Tickers 2026-04-30.xlsx", data_only=False)
trig = wb["Trig"]

print("Looking for derived indicator formulas in Trig sheet...\n")

# Map of rule names to search for
derived_rules = [
    'MACDH Direction',
    'MACD Direction',
    'BB Direction',
    'BBThresh Crossover',
    'Trade Cross Over',
    '!Trade Rule',
    'Trend Cross Over',
    '!Trend Rule',
    'Trend Trade Dep Rule',
    'Trade Trend Relation',
    '!Trade Trend Relation',
    'BRR% Dir Rule',
    'Trend below TRR',
    'LRR above Trade',
    'IVRule',
    '3mn Long Rule',
    '!Perf1D SD Rule',
    'Perf SD Rule',
    '!Perf SD Rule',
    '!Perf3D Rule',
    'BB Bull Rule',
    'BB Bull Puts',
    'MACD and H Rule',
    'MACD and H Rule Puts',
    '!Overbought',
    '!3wk Outlook',
    '!3wk Outlook Days',
    'Bull Rule',
    '!Bull Rule',
    'PerfOrBull Rule',
    '!PerfOrBull Rule',
    '50-DMA-Crossover',
    '200-DMA-Crossover',
    'Trade Close to BRR',
    'Trade Close to TRR',
    'Up Resistance',
    'Down Resistance',
    'VS LT Outlook Rule',
    'Short Term Oulook (If LT Bullish)',
    'Short Term Oulook (If LT Bearish)',
    'Overbought',
]

# Search through all rows for these rule names
for row_idx in range(1, trig.max_row + 1):
    cell_a = trig.cell(row_idx, 1).value

    if cell_a and str(cell_a).strip() in derived_rules:
        rule_name = str(cell_a).strip()

        # Print the rule and nearby cells to understand the formula
        print(f"\nRow {row_idx}: {rule_name}")

        # Print columns A-Q to see the structure
        for col_idx in range(1, 18):
            cell = trig.cell(row_idx, col_idx)
            col_letter = ''
            col = col_idx
            while col > 0:
                col -= 1
                col_letter = chr(65 + col % 26) + col_letter
                col //= 26

            if cell.value:
                print(f"  Col {col_letter}: {cell.value}")
