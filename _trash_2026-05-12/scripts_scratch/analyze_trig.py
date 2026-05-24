#!/usr/bin/env python
"""Analyze Trig sheet to understand derived indicator definitions."""
from openpyxl import load_workbook

wb = load_workbook("C:\\Ashok\\Invest\\Cluade\\Cluade\\Tickers 2026-04-30.xlsx", data_only=False)

print("Available sheets:")
for sheet in wb.sheetnames:
    print(f"  {sheet}")

# Look at the Trig sheet
if "Trig" in wb.sheetnames:
    trig = wb["Trig"]
    print(f"\n\nTrig sheet structure (first 120 rows):")

    # Find headers
    headers = []
    for col_idx in range(1, 70):
        cell = trig.cell(1, col_idx)
        if cell.value:
            headers.append((col_idx, cell.value))

    print(f"\nHeaders found in row 1:")
    for col_idx, header in headers[:30]:
        col_letter = chr(64 + col_idx) if col_idx <= 26 else chr(64 + col_idx // 26) + chr(64 + col_idx % 26)
        print(f"  Col {col_letter}: {header}")

    print(f"\n\nRule definitions (rows 4-120):")
    count = 0
    for row_idx in range(4, 121):
        rule_name = trig.cell(row_idx, 1).value
        ma_col_b = trig.cell(row_idx, 2).value
        ma_col_c = trig.cell(row_idx, 3).value
        from_val = trig.cell(row_idx, 4).value
        to_val = trig.cell(row_idx, 5).value

        if rule_name:
            count += 1
            print(f"  Row {row_idx}: {rule_name}")
            if ma_col_b:
                print(f"    Col B (MA?): {ma_col_b}")
            if ma_col_c:
                print(f"    Col C (MA?): {ma_col_c}")
            if from_val is not None:
                print(f"    From: {from_val}")
            if to_val is not None:
                print(f"    To: {to_val}")
            if count >= 45:
                break

    # Check columns O-EP as mentioned in CLAUDE.md for composite mapping
    print(f"\n\nColumns O onwards (composite rule mappings):")
    for col_idx in range(15, 35):  # O-AH
        col_letter = chr(64 + col_idx) if col_idx <= 26 else chr(64 + col_idx // 26) + chr(64 + col_idx % 26)
        header = trig.cell(1, col_idx).value
        if header:
            print(f"  Col {col_letter}: {header}")
