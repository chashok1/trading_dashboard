#!/usr/bin/env python
"""Extract formulas from Excel MA tab columns JG to NO."""
from openpyxl import load_workbook

wb = load_workbook('C:\Ashok\Invest\Projects\Cluade\Tickers 2026-04-30.xlsx', data_only=False)
ma = wb['MA']

def col_letter_to_num(col_str):
    num = 0
    for char in col_str:
        num = num * 26 + (ord(char) - ord('A') + 1)
    return num

def num_to_col_letter(col_idx):
    col_letter = ''
    col = col_idx
    while col > 0:
        col -= 1
        col_letter = chr(65 + col % 26) + col_letter
        col //= 26
    return col_letter

jg = col_letter_to_num('JG')  # 267
no = col_letter_to_num('NO')  # 379

print("Excel formulas for columns JG to NO:\n")
print("=" * 100)

# Get a data row to see formulas
data_row = 2  # First data row (row 1 is headers)

for col_idx in range(jg, no + 1):
    col_letter = num_to_col_letter(col_idx)
    header_cell = ma.cell(1, col_idx)
    formula_cell = ma.cell(data_row, col_idx)

    if header_cell.value:
        print(f"\n{col_letter}: {header_cell.value}")
        print(f"  Data Type: {formula_cell.data_type}")
        if formula_cell.value:
            if isinstance(formula_cell.value, str) and formula_cell.value.startswith('='):
                print(f"  Formula: {formula_cell.value}")
            else:
                print(f"  Value: {formula_cell.value}")
