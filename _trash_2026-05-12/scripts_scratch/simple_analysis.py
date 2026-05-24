#!/usr/bin/env python
from openpyxl import load_workbook

try:
    wb = load_workbook("C:\\Ashok\\Invest\\Projects\\Cluade\\Tickers 2026-05-06.xlsx", data_only=False)
    print("Workbook loaded successfully")
    print(f"Sheet names: {wb.sheetnames}")

    if "Trig" in wb.sheetnames:
        trig = wb["Trig"]
        print("\nTrig sheet found")
        print(f"Max row: {trig.max_row}")
        print(f"Max col: {trig.max_column}")

        # Print first 5 rows to see structure
        for row_idx in range(1, 6):
            row_vals = []
            for col_idx in range(1, 10):
                val = trig.cell(row_idx, col_idx).value
                row_vals.append(str(val)[:15] if val else "")
            print(f"Row {row_idx}: {row_vals}")
    else:
        print("Trig sheet not found")

except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
