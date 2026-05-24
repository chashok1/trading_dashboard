#!/usr/bin/env python
"""Analyze the Trig sheet to understand derived indicator mappings."""
from openpyxl import load_workbook

wb = load_workbook("C:\\Ashok\\Invest\\Projects\\Cluade\\Tickers 2026-05-06.xlsx", data_only=False)

if "Trig" in wb.sheetnames:
    trig = wb["Trig"]

    print("Trig sheet structure:\n")

    # Get header row (row 1)
    headers = []
    for col_idx in range(1, 25):
        cell = trig.cell(1, col_idx)
        if cell.value:
            col_letter = ''
            col = col_idx
            while col > 0:
                col -= 1
                col_letter = chr(65 + col % 26) + col_letter
                col //= 26
            headers.append((col_letter, cell.value))

    print("Headers in row 1:")
    for col_letter, header in headers:
        print(f"  Col {col_letter}: {header}")

    print("\n\nDerived Indicator Rules (rows 4-120, looking for those WITHOUT From/To):\n")

    for row_idx in range(4, 121):
        col_a = trig.cell(row_idx, 1).value  # Rule name
        col_b = trig.cell(row_idx, 2).value  # Name A?
        col_c = trig.cell(row_idx, 3).value  # Name B?
        col_d = trig.cell(row_idx, 4).value  # From
        col_e = trig.cell(row_idx, 5).value  # To

        if col_a and col_a.strip():
            # Check if this is a derived indicator (no From/To)
            if col_d is None and col_e is None:
                print(f"Row {row_idx}: {col_a}")
                if col_b:
                    print(f"  Col B: {col_b}")
                if col_c:
                    print(f"  Col C: {col_c}")
                # Print a few more columns to find the MA mapping
                for col_idx in range(6, 15):
                    val = trig.cell(row_idx, col_idx).value
                    if val:
                        col_letter = ''
                        col = col_idx
                        while col > 0:
                            col -= 1
                            col_letter = chr(65 + col % 26) + col_letter
                            col //= 26
                        print(f"  Col {col_letter}: {val}")
