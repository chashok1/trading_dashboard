#!/usr/bin/env python
"""Map all Excel MA tab columns to their meanings."""
from openpyxl import load_workbook

wb = load_workbook('C:\Ashok\Invest\Projects\Cluade\Tickers 2026-04-30.xlsx', data_only=False)
ma = wb['MA']

def num_to_col_letter(col_idx):
    col_letter = ''
    col = col_idx
    while col > 0:
        col -= 1
        col_letter = chr(65 + col % 26) + col_letter
        col //= 26
    return col_letter

# Get all column headers
headers = {}
for col_idx in range(1, ma.max_column + 1):
    header = ma.cell(1, col_idx).value
    col_letter = num_to_col_letter(col_idx)
    if header:
        headers[col_letter] = header

# Print columns referenced in the formulas we saw
referenced_cols = ['AN', 'AW', 'AX', 'AE', 'AF', 'EQ', 'ER', 'CI', 'CK', 'AC', 'BA', 'CH', 'CG',
                   'DX', 'DY', 'EH', 'EI', 'D', 'AC', 'MA', 'MB', 'MD', 'ME', 'LH', 'LK', 'LG',
                   'LW', 'MI', 'MM', 'MN', 'MQ', 'MS']

print("Excel column mappings (referenced in formulas):\n")
for col_letter in sorted(set(referenced_cols)):
    if col_letter in headers:
        print(f"  {col_letter}: {headers[col_letter]}")
    else:
        print(f"  {col_letter}: [NOT FOUND]")

print("\n\nAll MA tab columns:\n")
for col_idx in range(1, ma.max_column + 1):
    header = ma.cell(1, col_idx).value
    col_letter = num_to_col_letter(col_idx)
    if header:
        print(f"  {col_letter:4s}: {header}")
