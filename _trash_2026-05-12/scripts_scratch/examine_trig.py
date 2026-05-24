#!/usr/bin/env python
"""Examine Trig sheet structure in detail."""
from openpyxl import load_workbook

wb = load_workbook("C:\\Ashok\\Invest\\Projects\\Cluade\\Tickers 2026-04-30.xlsx", data_only=False)
trig = wb["Trig"]

print("Trig sheet structure:\n")
print(f"Max rows: {trig.max_row}")
print(f"Max cols: {trig.max_column}\n")

# Print first 5 rows with all columns
print("First 10 rows (all columns):")
for row_idx in range(1, 11):
    row_data = []
    for col_idx in range(1, 15):
        val = trig.cell(row_idx, col_idx).value
        row_data.append(str(val)[:12] if val else "-")
    print(f"Row {row_idx:2d}: {' | '.join(row_data)}")

print("\n\nFirst 20 data rows (rows 4-23):")
for row_idx in range(4, 24):
    col_a = trig.cell(row_idx, 1).value  # Rule name
    col_b = trig.cell(row_idx, 2).value  # ?
    col_c = trig.cell(row_idx, 3).value  # ?
    col_d = trig.cell(row_idx, 4).value  # From?
    col_e = trig.cell(row_idx, 5).value  # To?
    col_f = trig.cell(row_idx, 6).value

    print(f"Row {row_idx}: {str(col_a)[:30]:30s} | B:{str(col_b)[:15]:15s} | D:{str(col_d)[:8]:8s} | E:{str(col_e)[:8]:8s}")
