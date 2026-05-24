import openpyxl
import os
import tempfile

# Legacy columns to remove
legacy_cols = {
    'date_symbol',
    'date_symbol_time',
    'ds_key',
    'ds_sequence',
    'key_str',
    'sym_clean',
    'date_time_num'
}

# Load the workbook
wb = openpyxl.load_workbook('drv_formulas_reference.xlsx')
ws = wb.active

print(f'Total rows before: {ws.max_row}')
print(f'Processing sheet: {ws.title}')
print()

# Find and mark rows to delete (column B contains the column names)
rows_to_delete = []

for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
    # Check column B for column names
    cell_val = ws.cell(row_idx, 2).value
    if cell_val and isinstance(cell_val, str):
        col_name = cell_val.strip()
        if col_name in legacy_cols:
            rows_to_delete.append(row_idx)
            print(f'Row {row_idx}: Found legacy column "{col_name}"')

# Delete rows in reverse order to maintain indices
for row_idx in sorted(rows_to_delete, reverse=True):
    ws.delete_rows(row_idx, 1)

print()
print(f'Deleted {len(rows_to_delete)} rows')
print(f'Total rows after: {ws.max_row}')

# Save to temp file first, then replace original
temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
os.close(temp_fd)

try:
    wb.save(temp_path)
    # Replace original file by removing lock and old file first
    lock_file = '.~lock.drv_formulas_reference.xlsx#'
    if os.path.exists(lock_file):
        try:
            os.remove(lock_file)
        except:
            pass

    # Remove old file and rename temp
    if os.path.exists('drv_formulas_reference.xlsx'):
        os.remove('drv_formulas_reference.xlsx')
    os.rename(temp_path, 'drv_formulas_reference.xlsx')
    print()
    print('[OK] File saved successfully')
except Exception as e:
    print(f'[ERROR] Could not save: {e}')
    print('Make sure the file is not open in Excel')
finally:
    if os.path.exists(temp_path):
        try:
            os.remove(temp_path)
        except:
            pass
